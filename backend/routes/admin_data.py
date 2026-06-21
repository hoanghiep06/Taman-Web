# routes/admin_data.py
import io
import json
import pandas as pd
from fastapi import APIRouter, Depends, HTTPException, status, UploadFile, File
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from database import get_db
from models import User, Room, Elder, Asset, AuditLog
from core.dependencies import get_admin_user, get_privileged_user

router = APIRouter(prefix="/admin/data", tags=["Admin/Manager: Quản trị & Sao lưu dữ liệu"])

# =========================================================================
# ENDPOINT 1: DÀNH CHO CẢ ADMIN & MANAGER - NHẬP LIỆU MA TRẬN EXCEL
# =========================================================================
@router.post("/import-matrix", status_code=status.HTTP_201_CREATED)
async def import_matrix_assets_from_excel(
    file: UploadFile = File(...), 
    db: Session = Depends(get_db),
    current_user: User = Depends(get_privileged_user) # Cho phép cả Admin và Manager
):
    """
    Đọc file ma trận cấu trúc Excel (PHÒNG | TÊN CỤ | Điện thoại | Răng giả | ...)
    Tự động đồng bộ cấu trúc dữ liệu phòng, người cao tuổi và tài sản liên quan.
    """
    # KIỂM TRA CHẶT CHẼ: Đuôi file mở rộng
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST, 
            detail="Tệp tin không đúng định dạng. Hệ thống chỉ chấp nhận file Excel (.xlsx hoặc .xls)."
        )

    try:
        contents = await file.read()
        # Đọc dữ liệu, ép kiểu dữ liệu thô ban đầu tránh mất số phòng dạng "01", "02"
        df = pd.read_excel(io.BytesIO(contents), dtype={'PHÒNG': str, 'TÊN CỤ': str})
        
        # Chuẩn hóa tên các cột, xóa bỏ khoảng trắng thừa do người nhập gõ lỗi
        df.columns = [str(c).strip() for c in df.columns]
        
        # KIỂM TRA LOGIC HỆ THỐNG: Cấu trúc cột bắt buộc
        if 'PHÒNG' not in df.columns or 'TÊN CỤ' not in df.columns:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST, 
                detail="Cấu trúc file Excel sai biểu mẫu chuẩn. Bắt buộc phải có cột 'PHÒNG' và 'TÊN CỤ'."
            )

        # Xử lý các ô trộn dòng (Merged Cells) ở cột PHÒNG bằng kỹ thuật Forward Fill
        df['PHÒNG'] = df['PHÒNG'].astype(str).str.strip().replace(['nan', 'None', ''], None).ffill()

        # Trích xuất danh sách các cột tài sản vật phẩm phía sau
        asset_columns = [col for col in df.columns if col not in ['PHÒNG', 'TÊN CỤ']]

        rooms_created = 0
        elders_created = 0
        assets_created = 0

        # SỬ DỤNG TRANSACTION ATOMIC: Nếu một dòng lỗi, toàn bộ quá trình sẽ hủy bỏ (All or Nothing)
        with db.begin_nested():
            for index, row in df.iterrows():
                # Bỏ qua các dòng trống hoàn toàn hoặc dữ liệu rác
                if pd.isna(row['PHÒNG']) or pd.isna(row['TÊN CỤ']):
                    continue
                
                room_num = str(row['PHÒNG']).strip()
                elder_name = str(row['TÊN CỤ']).strip()

                if not room_num or not elder_name or room_num.lower() == 'nan' or elder_name.lower() == 'nan':
                    continue

                # LOGIC 1: Xử lý đồng bộ Phòng
                room = db.query(Room).filter(Room.room_number == room_num).first()
                if not room:
                    room = Room(room_number=room_num, description=f"Phòng {room_num} khởi tạo tự động từ file Excel")
                    db.add(room)
                    db.flush() # Lấy ID phòng ngay lập tức
                    rooms_created += 1

                # LOGIC 2: Xử lý đồng bộ Người Cao Tuổi trong phòng đó
                elder = db.query(Elder).filter(Elder.full_name == elder_name, Elder.room_id == room.id).first()
                if not elder:
                    elder = Elder(full_name=elder_name, room_id=room.id)
                    db.add(elder)
                    db.flush() # Lấy ID cụ ngay lập tức
                    elders_created += 1

                # LOGIC 3: Quét ma trận bóc tách dấu tích vật phẩm
                for asset_name in asset_columns:
                    cell_value = row[asset_name]
                    
                    # Bộ lọc nhận diện dấu tích cực kỳ thông minh (chấp nhận dấu x, X, số 1, dấu check v, ✓)
                    is_checked = False
                    if pd.notna(cell_value):
                        val_str = str(cell_value).strip().lower()
                        if val_str in ['1', '1.0', 'true', 'x', 'v', '✓', 'yes']:
                            is_checked = True

                    if is_checked:
                        # Kiểm tra trùng lặp tài sản để tránh ghi đè dữ liệu cũ đang có
                        existing_asset = db.query(Asset).filter(
                            Asset.asset_name == asset_name,
                            Asset.room_id == room.id,
                            Asset.elder_id == elder.id
                        ).first()

                        if not existing_asset:
                            new_asset = Asset(
                                asset_name=asset_name,
                                room_id=room.id,
                                elder_id=elder.id,
                                status='Active'
                            )
                            db.add(new_asset)
                            assets_created += 1

        db.commit()
        return {
            "status": "Success",
            "message": f"Tài khoản [{current_user.username}] đã nạp dữ liệu ma trận thành công!",
            "summary": {
                "phòng_mới_bổ_sung": rooms_created,
                "người_cao_tuổi_mới_bổ_sung": elders_created,
                "vật_phẩm_kích_hoạt": assets_created
            }
        }

    except HTTPException as http_err:
        db.rollback()
        raise http_err
    except Exception as e:
        db.rollback()
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, 
            detail=f"Cấu trúc tệp dữ liệu không đồng nhất hoặc bị lỗi hàng hàng loạt. Chi tiết: {str(e)}"
        )


# =========================================================================
# ENDPOINT 2: CHỈ DÀNH CHO ADMIN - XUẤT ĐA DẠNG TAB ĐỂ BACKUP/BẢO TRÌ SYSTEM
# =========================================================================
@router.get("/backup", response_class=StreamingResponse)
def backup_and_audit_database(
    db: Session = Depends(get_db),
    current_admin: User = Depends(get_admin_user) # KHÓA CHẶT: Chỉ riêng Admin tối cao
):
    """
    Hàm xuất dữ liệu tối ưu: Gom toàn bộ dữ liệu thực thể và lịch sử log an ninh 
    vào một file Excel đa tầng được định dạng chuyên nghiệp phục vụ bảo trì.
    """
    # 1. Khởi tạo cây dữ liệu tổng hợp (Master Sheet) dạng phẳng từ các phép Join DB
    master_query = db.query(
        Room.room_number, Elder.full_name, Asset.asset_name, Asset.status, Asset.created_at
    ).join(Elder, Asset.elder_id == Elder.id).join(Room, Asset.room_id == Room.id).all()

    master_data = [{
        "Số Phòng": q[0], "Tên Người Cao Tuổi": q[1], "Tên Tài Sản/Vật Phẩm": q[2], 
        "Trạng Thái Hoạt Động": q[3], "Ngày Đưa Vào Sử Dụng": q[4].strftime("%d/%m/%Y") if q[4] else ""
    } for q in master_query]

    # 2. Thu thập dữ liệu các bảng gốc hỗ trợ khôi phục khi sập hệ thống
    rooms = db.query(Room).all()
    elders = db.query(Elder).all()
    audit_logs = db.query(AuditLog).order_by(AuditLog.created_at.desc()).limit(100).all() # Lấy 100 log mới nhất

    room_data = [{"ID Phòng": r.id, "Số Phòng": r.room_number, "Mô Tả": r.description} for r in rooms]
    elder_data = [{"ID Cụ": e.id, "Tên Cụ Già": e.full_name, "Thuộc ID Phòng": e.room_id} for e in elders]
    
    audit_data = [{
        "Mã Log": log.id, "Mã Nhân Viên": log.actor_id, "Hành Động": log.action,
        "Thực Thể Đích": log.target_id, "Địa Chỉ IP": log.ip_address, 
        "Thời Gian Hệ Thống": log.created_at.strftime("%d/%m/%Y %H:%M:%S") if log.created_at else ""
    } for log in audit_logs]

    # 3. Chuyển đổi dữ liệu sang DataFrames của Pandas
    df_master = pd.DataFrame(master_data)
    df_rooms = pd.DataFrame(room_data)
    df_elders = pd.DataFrame(elder_data)
    df_audit = pd.DataFrame(audit_data)

    # 4. Sử dụng BytesIO để ghi dữ liệu trực tiếp trên RAM, xuất luồng siêu tốc
    output = io.BytesIO()
    
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df_master.to_excel(writer, sheet_name='MASTER_MA_TRAN', index=False)
        df_rooms.to_excel(writer, sheet_name='DANH_MUC_PHONG', index=False)
        df_elders.to_excel(writer, sheet_name='DANH_SACH_NCT', index=False)
        df_audit.to_excel(writer, sheet_name='LOGS_AN_NINH_H_T', index=False)

        # ---- LỚP TRANG TRÍ ĐỊNH DẠNG EXCEL CHUYÊN NGHIỆP TRỰC TRỰC TIẾP TRÊN FILE XUẤT ----
        workbook = writer.book
        # Bảng màu chuẩn doanh nghiệp
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid") # Màu xanh nước biển đậm
        header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
        center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style='thin', color='D9D9D9'), right=Side(style='thin', color='D9D9D9'),
            top=Side(style='thin', color='D9D9D9'), bottom=Side(style='thin', color='D9D9D9')
        )

        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            # Định dạng hàng đầu tiên (Header)
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_align
            
            # Cấu hình tự động căn rộng lề ô (Auto-fit Column Widths) theo độ dài chữ thực tế
            for col in worksheet.columns:
                max_len = max(len(str(cell.value or '')) for cell in col)
                col_letter = col[0].column_letter
                worksheet.column_dimensions[col_letter].width = max(max_len + 4, 12)
                
                # Đóng khung mờ cho toàn bộ ô dữ liệu giúp xem không bị rối mắt
                for cell in col:
                    if cell.row > 1:
                        cell.border = thin_border
                        cell.font = Font(name="Arial", size=10)

    output.seek(0)
    timestamp = pd.Timestamp.now().strftime("%Y%m%d_%H%M%S")
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=TamAn_SystemBackup_{timestamp}.xlsx"}
    )