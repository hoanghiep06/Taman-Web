# routes/admin_entities.py
import io
import os
import logging
import threading
import models
from datetime import datetime, timedelta, timezone
from typing import List, Optional
from fastapi import UploadFile, File, APIRouter, Depends, HTTPException, status, Query
from sqlalchemy.orm import Session
from sqlalchemy import text
from openpyxl import load_workbook
from database import get_db, engine
from models import User, Room, Elder, Asset
from schemas import ElderCreate, ElderResponse, AssetCreate, AssetResponse
# Nạp cả get_current_user (mọi tài khoản) và get_privileged_user (chức vụ cao)
from core.dependencies import get_privileged_user, get_current_user, get_admin_user
from services.backup_service import execute_database_dump, execute_database_restore
from services.drive_service import (
    upload_db_backup_to_drive,
    download_file_bytes_from_drive,
    list_db_backups_from_drive,
    cleanup_old_db_backups,
)
from core.constants import DEFAULT_SHIFT_SETTINGS

logger = logging.getLogger("disaster_recovery")
# Khóa chống chạy khôi phục đồng thời: tránh 2 admin cùng bấm restore 1 lúc gây xung đột DB
_restore_lock = threading.Lock()

# ==========================================
# KHAI BÁO CÁC ROUTER RIÊNG BIỆT (GỠ BỎ DEPENDENCIES CẤP TỔNG)
# ==========================================
router_elders = APIRouter(
    prefix="/admin/elders", 
    tags=["Quản lý Người Cao Tuổi (NCT)"]
)

router_assets = APIRouter(
    prefix="/admin/assets", 
    tags=["Quản lý Tài Sản"]
)

router_backup = APIRouter(
    prefix="/admin/system/backup",
      tags=["Admin: Backup dữ liệu"])


# =========================================================================
# PHÂN HỆ NGHIỆP VỤ: QUẢN LÝ NGƯỜI CAO TUỔI (NCT)
# =========================================================================

# 1. API THÊM CỤ GIÀ MỚI (CHỈ ADMIN/MANAGER ĐƯỢC PHÉP)
@router_elders.post("", response_model=ElderResponse, status_code=status.HTTP_201_CREATED)
def create_elder(
    elder: ElderCreate, 
    db: Session = Depends(get_db),
    current_user: User = Depends(get_privileged_user) # Khóa chặt chức năng ghi
):
    if elder.room_id:
        room = db.query(Room).filter(Room.id == elder.room_id).first()
        if not room:
            raise HTTPException(status_code=404, detail="Không tìm thấy phòng này")
            
    new_elder = Elder(**elder.model_dump())
    db.add(new_elder)
    db.commit()
    db.refresh(new_elder)
    return new_elder


# 2. API LẤY TOÀN BỘ DANH SÁCH CỤ GIÀ (MỞ RỘNG CHO AI CŨNG LẤY ĐƯỢC - KỂ CẢ STAFF)
@router_elders.get("", response_model=List[ElderResponse])
def get_all_elders(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user) # ĐÃ SỬA: Bất kỳ ai đăng nhập đều đọc được
):
    """
    Trả về danh sách toàn bộ các cụ trong trung tâm giúp ứng dụng di động 
    của nhân viên vẽ cấu trúc ma trận phòng ốc trực quan.
    """
    return db.query(Elder).all()


# 3. API LẤY CHI TIẾT 1 CỤ GIÀ (MỞ RỘNG CHO AI CŨNG LẤY ĐƯỢC - KỂ CẢ STAFF)
@router_elders.get("/{elder_id}", response_model=ElderResponse)
def get_elder(
    elder_id: int, 
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user) # ĐÃ SỬA: Bất kỳ ai đăng nhập đều đọc được
):
    elder = db.query(Elder).filter(Elder.id == elder_id).first()
    if not elder:
        raise HTTPException(status_code=404, detail="Không tìm thấy thông tin cụ")
    return elder


# 4. API SỬA THÔNG TIN CỤ GIÀ (CHỈ ADMIN/MANAGER ĐƯỢC PHÉP)
@router_elders.put("/{elder_id}", response_model=ElderResponse)
def update_elder(
    elder_id: int, 
    elder_data: ElderCreate, 
    db: Session = Depends(get_db),
    current_user: User = Depends(get_privileged_user)
):
    elder = db.query(Elder).filter(Elder.id == elder_id).first()
    if not elder:
        raise HTTPException(status_code=404, detail="Không tìm thấy thông tin cụ")

    # Lưu lại ID phòng cũ để đối chiếu
    old_room_id = elder.room_id

    if elder_data.room_id:
        room = db.query(Room).filter(Room.id == elder_data.room_id).first()
        if not room:
            raise HTTPException(status_code=404, detail="Không tìm thấy phòng này")

    for field, value in elder_data.model_dump(exclude_unset=True).items():
        setattr(elder, field, value)

    # ──── ĐOẠN VÁ LỖI ĐỒNG BỘ: TỰ ĐỘNG DI DỜI TÀI SẢN KHI CỤ ĐỔI PHÒNG ────
    if elder.room_id != old_room_id:
        db.query(Asset).filter(Asset.elder_id == elder.id).update(
            {"room_id": elder.room_id}, 
            synchronize_session=False
        )

    db.commit()
    db.refresh(elder)
    return elder

# 5. API XÓA CỤ GIÀ (CHỈ ADMIN/MANAGER ĐƯỢC PHÉP)
@router_elders.delete("/{elder_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_elder(
    elder_id: int, 
    db: Session = Depends(get_db),
    current_user: User = Depends(get_privileged_user) # Khóa chặt chức năng ghi
):
    elder = db.query(Elder).filter(Elder.id == elder_id).first()
    if not elder:
        raise HTTPException(status_code=404, detail="Không tìm thấy thông tin cụ")

    db.delete(elder)
    db.commit()


# =========================================================================
# PHÂN HỆ NGHIỆP VỤ: QUẢN LÝ DANH MỤC TÀI SẢN
# =========================================================================

# 1. API TẠO TÀI SẢN (CHỈ ADMIN/MANAGER ĐƯỢC PHÉP)
@router_assets.post("", response_model=AssetResponse, status_code=status.HTTP_201_CREATED)
def create_asset(
    asset: AssetCreate, 
    db: Session = Depends(get_db),
    current_user: User = Depends(get_privileged_user)
):
    asset_data = asset.model_dump()

    # ──── ĐOẠN VÁ LỖI: ÉP BUỘC ĐỒNG BỘ PHÒNG THEO CỤ GIÀ KHI TẠO MỚI ────
    if asset.elder_id:
        elder = db.query(Elder).filter(Elder.id == asset.elder_id).first()
        if not elder:
            raise HTTPException(status_code=404, detail="Không tìm thấy thông tin cụ già này")
        # Điền đè phòng của tài sản trùng khít với phòng của cụ quản lý
        asset_data["room_id"] = elder.room_id
        
    new_asset = Asset(**asset_data)
    db.add(new_asset)
    db.commit()
    db.refresh(new_asset)
    return new_asset


# 2. API XEM DANH SÁCH TÀI SẢN (MỞ RỘNG CHO AI CŨNG LẤY ĐƯỢC - KỂ CẢ STAFF)
@router_assets.get("", response_model=List[AssetResponse])
def get_all_assets(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user) # ĐÃ SỬA: Bất kỳ ai đăng nhập đều đọc được
):
    return db.query(Asset).all()


# 3. API XEM CHI TIẾT TÀI SẢN (MỞ RỘNG CHO AI CŨNG LẤY ĐƯỢC - KỂ CẢ STAFF)
@router_assets.get("/{asset_id}", response_model=AssetResponse)
def get_asset(
    asset_id: int, 
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user) # ĐÃ SỬA: Bất kỳ ai đăng nhập đều đọc được
):
    asset = db.query(Asset).filter(Asset.id == asset_id).first()
    if not asset:
        raise HTTPException(status_code=404, detail="Không tìm thấy tài sản này")
    return asset


# 4. API CẬP NHẬT TÀI SẢN (CHỈ ADMIN/MANAGER ĐƯỢC PHÉP)
@router_assets.put("/{asset_id}", response_model=AssetResponse)
def update_asset(
    asset_id: int, 
    asset_data: AssetCreate, 
    db: Session = Depends(get_db),
    current_user: User = Depends(get_privileged_user)
):
    asset = db.query(Asset).filter(Asset.id == asset_id).first()
    if not asset:
        raise HTTPException(status_code=404, detail="Không tìm thấy tài sản này")

    for field, value in asset_data.model_dump(exclude_unset=True).items():
        setattr(asset, field, value)

    # ──── ĐOẠN VÁ LỖI: ĐỒNG BỘ PHÒNG CỦA TÀI SẢN THEO CHỦ SỞ HỮU MỚI ────
    if asset.elder_id:
        elder = db.query(Elder).filter(Elder.id == asset.elder_id).first()
        if elder:
            asset.room_id = elder.room_id

    db.commit()
    db.refresh(asset)
    return asset


# 5. API XÓA TÀI SẢN (CHỈ ADMIN/MANAGER ĐƯỢC PHÉP)
@router_assets.delete("/{asset_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_asset(
    asset_id: int, 
    db: Session = Depends(get_db),
    current_user: User = Depends(get_privileged_user) # Khóa chặt chức năng ghi
):
    asset = db.query(Asset).filter(Asset.id == asset_id).first()
    if not asset:
        raise HTTPException(status_code=404, detail="Không tìm thấy tài sản này")

    db.delete(asset)
    db.commit()



@router_assets.post("/import-xlsx", status_code=status.HTTP_200_OK)
async def import_assets_from_xlsx(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_privileged_user)
):
    """
    API Import danh sách vật tư theo cấu trúc Ma trận chuẩn của trung tâm Tâm An:
    - Cột A (Index 0): Số phòng (Có thể để trống nếu ở chung phòng với cụ dòng trên)
    - Cột B (Index 1): Họ và tên cụ
    - Cột C trở đi: Ô giá trị TRUE/FALSE ứng với Tên món đồ nằm ở dòng Tiêu đề 1.
    """
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST, 
            detail="Hệ thống chỉ chấp nhận file Excel định dạng .xlsx hoặc .xls"
        )

    try:
        # 1. Nạp tệp nhị phân vào bộ nhớ RAM
        file_contents = await file.read()
        wb = load_workbook(io.BytesIO(file_contents), data_only=True)
        ws = wb.active 
        
        # 2. Bóc tách danh sách tên món đồ từ Dòng tiêu đề số 1 (Header Row)
        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
        headers = [str(cell).strip() if cell else "" for cell in header_row]

        processed_elders = 0
        total_assets_upserted = 0
        
        # Biến nhớ trung gian gác cổng để lưu số phòng của cụ dòng trên
        current_room_number = None

        # 3. Quét ma trận từ dòng dữ liệu số 2 trở đi
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or len(row) < 2:
                continue

            room_raw = row[0]   # Cột A: PHÒNG
            elder_raw = row[1]  # Cột B: TÊN CỤ

            # KỸ THUẬT STICKY ROOM: Nếu dòng này có số phòng mới -> Cập nhật; Nếu trống -> Thừa kế phòng dòng trên
            if room_raw is not None and str(room_raw).strip() != "":
                current_room_number = str(room_raw).strip()

            # Nếu tên cụ trống hoặc file bị lỗi không bắt được số phòng ban đầu -> Bỏ qua dòng
            if not current_room_number or elder_raw is None or str(elder_raw).strip() == "":
                continue

            elder_name = str(elder_raw).strip()

            # ──── LOGIC TẦNG 1: XỬ LÝ PHÒNG ────
            room = db.query(Room).filter(Room.room_number == current_room_number).first()
            if not room:
                room = Room(room_number=current_room_number, description=f"Phòng {current_room_number}")
                db.add(room)
                db.flush() 

            # ──── LOGIC TẦNG 2: XỬ LÝ CỤ GIÀ ────
            elder = db.query(Elder).filter(
                Elder.full_name == elder_name, 
                Elder.room_id == room.id
            ).first()
            
            if not elder:
                elder = Elder(full_name=elder_name, room_id=room.id)
                db.add(elder)
                db.flush()
            
            processed_elders += 1

            # ──── LOGIC TẦNG 3: BẢO TRÌ DỮ LIỆU CŨ (SOFT REPLACE) ────
            existing_assets = db.query(Asset).filter(Asset.elder_id == elder.id).all()
            asset_dict = {a.asset_name.lower().strip(): a for a in existing_assets}
            
            # Đưa toàn bộ trạng thái đồ đạc cũ về kho lưu trữ ẩn (Archived) để thỏa mãn Constraint
            for asset in existing_assets:
                asset.status = "Archived"

            # ──── LOGIC TẦNG 4: THUẬT TOÁN QUÉT DÀN NGANG THEO TIÊU ĐỀ ────
            # Duyệt từ cột index 2 (Cột C trở đi) ứng với các món đồ vật tư trong file
            for col_index in range(2, len(row)):
                if col_index >= len(headers):
                    break

                asset_name = headers[col_index]
                # Bỏ qua nếu cột tiêu đề bị trống hoặc vô tình trùng với cột gác cổng
                if not asset_name or asset_name in ["PHÒNG", "TÊN CỤ"]:
                    continue

                cell_val = row[col_index]
                # openpyxl tự nhận diện kiểu dữ liệu Boolean hoặc so khớp chuỗi chữ hoa "TRUE"
                is_checked = (cell_val is True) or (str(cell_val).strip().upper() == "TRUE")

                if is_checked:
                    search_key = asset_name.lower().strip()
                    
                    if search_key in asset_dict:
                        # Đồ vật đã tồn tại -> Bật kích hoạt Active trở lại
                        asset_dict[search_key].status = "Active"
                        # ĐBỔ SUNG: Đồng bộ luôn phòng mới trong trường hợp cụ đổi phòng
                        asset_dict[search_key].room_id = room.id
                    else:
                        # Món đồ mới tinh chưa từng có -> Thêm mới vào DB
                        new_asset = Asset(
                            asset_name=asset_name,
                            room_id=room.id,
                            elder_id=elder.id,
                            status="Active"
                        )
                        db.add(new_asset)
                        
                    total_assets_upserted += 1

        db.commit()

        return {
            "status": "Success",
            "message": "Đồng bộ ma trận danh mục vật tư từ file Excel thành công mỹ mãn!",
            "summary": {
                "total_elders_processed": processed_elders,
                "total_assets_active_count": total_assets_upserted
            }
        }

    except Exception as e:
        db.rollback()
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, 
            detail=f"Lỗi ma trận tệp Excel: {str(e)}"
        )
    




@router_backup.get("/list", status_code=status.HTTP_200_OK)
def get_backup_list_on_cloud(
    current_user: User = Depends(get_admin_user)
):
    """ 
    API đặc quyền giúp Admin xem danh sách toàn bộ các file backup hiện có trên Drive,
    bao gồm File ID, Tên file, Kích thước (bytes) và Thời gian khởi tạo để làm UI chọn file khôi phục.
    """
    try:
        backups = list_db_backups_from_drive()
        return {
            "status": "Success",
            "total_backups_found": len(backups),
            "backups": backups
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Không thể truy vấn danh sách Cloud: {str(e)}")


@router_backup.post("/manual-run", status_code=status.HTTP_200_OK)
def trigger_manual_backup(
    current_user: User = Depends(get_admin_user)
):
    """
    API giúp Admin chủ động bấm nút "Sao lưu ngay lập tức" trước khi thực hiện các thao tác
    bảo trì hệ thống nặng, tự động đẩy lên /backup và chạy luôn bộ lọc giữ đúng 4 bản.
    """
    logger.info(f"[BACKUP] User '{current_user.username}' (id={current_user.id}) kích hoạt backup thủ công.")
    try:
        sql_bytes = execute_database_dump()
        timestamp_str = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"TamAn_Manual_DB_Backup_{timestamp_str}.sql"
        
        drive_link = upload_db_backup_to_drive(file_bytes=sql_bytes, filename=filename)
        
        # Ép dọn dẹp luôn sau khi sao lưu tay
        cleanup_old_db_backups(keep_count=4)
        
        logger.info(f"[BACKUP SUCCESS] {filename} -> {drive_link}")
        return {
            "status": "Success",
            "message": "Sao lưu thủ công cấp tốc thành công!",
            "drive_url": drive_link
        }
    except Exception as e:
        logger.error(f"[BACKUP FAILED] {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@router_backup.post("/restore", status_code=status.HTTP_200_OK)
def restore_database_system(
    drive_file_id: Optional[str] = Query(None, description="Truyền ID file trên Google Drive để khôi phục trực tiếp"),
    file: Optional[UploadFile] = File(None, description="Hoặc đính kèm file .sql từ máy tính cá nhân lên"),
    db: Session = Depends(get_db), # 🔥 BỔ SUNG: Lấy Session db hiện tại của Request vào đây
    current_user: User = Depends(get_admin_user)
):
    """
    API CỨU HỘ TỐI CAO (Disaster Recovery)
    """
    if not _restore_lock.acquire(blocking=False):
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail="Đang có một tiến trình khôi phục khác chạy, vui lòng chờ rồi thử lại."
        )

    safety_link = None
    try:
        sql_data_bytes = b""

        # Tình huống 1: Khôi phục trực tiếp bằng file ID của Google Drive
        if drive_file_id:
            logger.warning(f"[RECOVERY] User '{current_user.username}' (id={current_user.id}) khởi tạo restore từ Drive file_id='{drive_file_id}'")
            sql_data_bytes = download_file_bytes_from_drive(drive_file_id)

        # Tình huống 2: Khôi phục bằng file đính kèm nộp từ Client
        elif file:
            if not file.filename.endswith('.sql'):
                raise HTTPException(status_code=400, detail="Hệ thống chỉ chấp nhận tệp tin có định dạng .sql")
            logger.warning(f"[RECOVERY] User '{current_user.username}' (id={current_user.id}) khởi tạo restore từ file upload '{file.filename}'")
            sql_data_bytes = file.file.read()

        else:
            raise HTTPException(
                status_code=400,
                detail="Vui lòng cung cấp ít nhất một phương thức khôi phục: Truyền 'drive_file_id' hoặc Nộp tệp tin 'file'."
            )

        if not sql_data_bytes:
            raise HTTPException(status_code=400, detail="Dữ liệu file phục hồi rỗng hoặc bị lỗi.")

        # ──── ĐOẠN VÁ LỖI XUNG ĐỘT PHIÊN BẢN (COMPATIBILITY SHIELD) ────
        if b"SET transaction_timeout = 0;" in sql_data_bytes:
            logger.warning("[RECOVERY]: Phát hiện chỉ thị 'transaction_timeout' của PG17. Tiến hành vô hiệu hóa để tương thích với cấu trúc PG15 hiện tại...")
            sql_data_bytes = sql_data_bytes.replace(
                b"SET transaction_timeout = 0;", 
                b"-- SET transaction_timeout = 0;"
            )

        # ──── BƯỚC AN TOÀN BẮT BUỘC: snapshot DB hiện tại trước khi ghi đè ────
        try:
            logger.warning("[RECOVERY] Đang tạo snapshot an toàn của DB hiện tại trước khi ghi đè...")
            safety_bytes = execute_database_dump()
            ts = datetime.now().strftime("%Y%m%d_%H%M%S")
            safety_filename = f"TamAn_PreRestoreSafety_{ts}.sql"
            safety_link = upload_db_backup_to_drive(file_bytes=safety_bytes, filename=safety_filename)
            logger.warning(f"[RECOVERY] Đã lưu snapshot an toàn lên Drive: {safety_link}")
        except Exception as safety_err:
            logger.error(f"[RECOVERY ABORTED] Không tạo được snapshot an toàn nên đã HỦY restore: {safety_err}")
            raise HTTPException(
                status_code=500,
                detail=f"Đã hủy khôi phục vì không thể tạo bản backup an toàn trước khi ghi đè: {safety_err}"
            )

        # =========================================================================
        # 🔥 BƯỚC CHÍ MẠNG: CHỦ ĐỘNG GIẢI PHÓNG KẾT NỐI ĐỂ BẺ GÃY DEADLOCK
        # =========================================================================
        # Đóng session hiện tại và dọn sạch Connection Pool của FastAPI. 
        # Hành động này giải phóng hoàn toàn các AccessShareLock trên DB live, 
        # nhường đường cho lệnh khôi phục thô bên dưới chiếm quyền ghi exclusive mượt mà!
        db.close()
        from database import engine
        engine.dispose()

        # Kích hoạt lệnh nạp ngầm nhị phân psql tái cấu trúc máy chủ (chạy trong 1 transaction)
        execute_database_restore(sql_data_bytes)

        from sqlalchemy.orm import sessionmaker
        from services.shift_service import auto_open_shift # Nạp hàm mở ca trực
        
        FreshSession = sessionmaker(bind=engine)
        fallback_db = FreshSession()

        try:
            vietnam_tz = timezone(timedelta(hours=7))
            current_hour = datetime.now(vietnam_tz).hour
            logger.warning(f"[RECOVERY]: Đang tiến hành rà soát khung giờ ({current_hour}h) để khôi phục ca trực live sau restore...")

            # Truy vấn khung giờ cấu hình thực tế từ bảng shift_settings vừa khôi phục thành công từ file SQL
            setting = fallback_db.query(models.ShiftSetting).first()
            if setting:
                m_start = int(setting.morning_start.split(':')[0])
                m_end = int(setting.morning_end.split(':')[0])
                e_start = int(setting.evening_start.split(':')[0])
                e_end = int(setting.evening_end.split(':')[0])
            else:
                m_start = int(DEFAULT_SHIFT_SETTINGS["morning_start"].split(':')[0])
                m_end = int(DEFAULT_SHIFT_SETTINGS["morning_end"].split(':')[0])
                e_start = int(DEFAULT_SHIFT_SETTINGS["evening_start"].split(':')[0])
                e_end = int(DEFAULT_SHIFT_SETTINGS["evening_end"].split(':')[0])

            # Đối chiếu số nguyên an toàn để sinh bù ca trực nếu file backup cũ bị thiếu
            if m_start <= current_hour < m_end:
                auto_open_shift(fallback_db, "Sang")
                logger.info("[RECOVERY SUCCESS]: Đã tự động bù đắp lại Ca Sáng thành công cho ngày hôm nay!")
            elif e_start <= current_hour < e_end:
                auto_open_shift(fallback_db, "Toi")
                logger.info("[RECOVERY SUCCESS]: Đã tự động bù đắp lại Ca Tối thành công cho ngày hôm nay!")
            else:
                logger.info("[RECOVERY INFO]: Thời điểm khôi phục nằm ngoài khung giờ hành chính, nhường quyền cho Scheduler tự động sinh ca tiếp theo.")

        except Exception as recovery_shift_err:
            fallback_db.rollback()
            logger.error(f"[RECOVERY SHIFT WARN]: Quá trình tự động bù đắp ca trực live gặp sự cố: {str(recovery_shift_err)}")
        finally:
            fallback_db.close() # Giải phóng kết nối an toàn

        logger.warning(f"[RECOVERY SUCCESS] User '{current_user.username}' đã khôi phục DB thành công.")
        return {
            "status": "Success",
            "message": "Hệ thống đã được khôi phục toàn vẹn nguyên bản thành công mỹ mãn!",
            "pre_restore_safety_backup_url": safety_link
        }

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"[RECOVERY FATAL_ERROR] {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Quy trình cứu hộ thảm họa thất bại: {str(e)}"
        )
    finally:
        _restore_lock.release()



@router_backup.post("/reset-database", status_code=status.HTTP_200_OK)
def hard_reset_database_system(
    db: Session = Depends(get_db), 
    current_user: User = Depends(get_admin_user)
):
    """
    API SIÊU KHẨN CẤP (Hard Reset System) - CHỈ DÀNH CHO ADMIN TỐI CAO
    """
    if not _restore_lock.acquire(blocking=False):
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail="Hệ thống đang bận xử lý một tác vụ cứu hộ dữ liệu khác."
        )

    try:
        # BƯỚC 1: SAO LƯU AN TOÀN TRƯỚC KHI PHÁ HỦY DỮ LIỆU CŨ
        logger.warning(f"[HARD RESET DB]: Admin '{current_user.username}' kích hoạt lệnh xóa trắng DB live. Đang làm snapshot dự phòng...")
        try:
            pre_reset_bytes = execute_database_dump()
            ts = datetime.now().strftime("%Y%m%d_%H%M%S") # Sử dụng mượt mà từ import đầu file
            safety_filename = f"TamAn_PreHardResetSafety_{ts}.sql"
            safety_drive_link = upload_db_backup_to_drive(file_bytes=pre_reset_bytes, filename=safety_filename)
            logger.warning(f"[HARD RESET DB]: Đã lưu snapshot dự phòng thành công tại: {safety_drive_link}")
        except Exception as backup_err:
            logger.error(f"[HARD RESET ABORTED]: Không thể sao lưu dự phòng, HỦY lệnh xóa trắng DB: {backup_err}")
            raise HTTPException(
                status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                detail=f"Lệnh xóa trắng DB bị hủy bỏ để bảo vệ an toàn: {backup_err}"
            )

        # Đóng session hiện tại để giải phóng trạng thái kết nối cũ
        db.close()
        from database import engine

        # =========================================================================
        # BƯỚC 2: GỘT RỬA VÀ TÁI THIẾT LẬP DB BẰNG ORM (TRÁNH LỖI SẬP KẾT NỐI)
        # =========================================================================
        logger.warning("[HARD RESET DB]: Tiến hành gỡ bỏ toàn bộ cấu trúc bảng cũ qua ORM...")
        models.Base.metadata.drop_all(bind=engine)

        logger.warning("[HARD RESET DB]: Tiến hành tái thiết lập cấu trúc bảng sạch mới tinh...")
        models.Base.metadata.create_all(bind=engine)

        # =========================================================================
        # BƯỚC 3: TÁI TẠO PHÂN HỆ CHỈ MỤC INDEXES ĐỂ ĐẢM BẢO TỐC ĐỘ TRA CỨU
        # =========================================================================
        logger.warning("[HARD RESET DB]: Đang thiết lập các chỉ mục Indexes tối ưu...")
        with engine.connect() as connection:
            connection.execute(text("CREATE INDEX idx_inspection_logs_latest ON inspection_logs (shift_id, asset_id) WHERE is_latest = TRUE;"))
            connection.execute(text("CREATE INDEX idx_assets_room ON assets (room_id);"))
            connection.execute(text("CREATE INDEX idx_shift_summaries_created ON shift_summaries (created_at DESC);"))
            connection.commit()

        # =========================================================================
        # BƯỚC 4: SEED ACCOUNT ADMIN ĐỘNG VỚI MẬT KHẨU BĂM CHUẨN 100% & MỞ LẠI CA LIVE
        # =========================================================================
        from sqlalchemy.orm import sessionmaker
        FreshSession = sessionmaker(bind=engine)
        fallback_db = FreshSession()

        try:
            logger.warning("[HARD RESET DB]: Tiến hành nạp tài khoản tối cao mặc định...")
            from passlib.context import CryptContext
            pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
            secure_hashed_password = pwd_context.hash("123456")

            admin_user = User(
                username="admin",
                password_hash=secure_hashed_password,
                full_name="Quản Trị Viên Hệ Thống",
                role="Admin",
                is_active=True,
                must_change_password=True
            )
            fallback_db.add(admin_user)
            fallback_db.commit()
            logger.info("[HARD RESET SUCCESS]: Tài khoản admin khởi tạo thành công mượt mà!")

            # ──── KHÔI PHỤC CA TRỰC LIVE SỬ DỤNG ÉP KIỂU SỐ NGUYÊN AN TOÀN ────
            from services.shift_service import auto_open_shift
            
            vietnam_tz = timezone(timedelta(hours=7))
            current_hour = datetime.now(vietnam_tz).hour
            logger.warning(f"[HARD RESET DB]: Đang kiểm tra khung giờ hiện tại ({current_hour}h) để khôi phục ca trực live...")

            # Bóc tách khung giờ từ Database live để kiểm tra chính xác
            setting = fallback_db.query(models.ShiftSetting).first()
            if setting:
                m_start = int(setting.morning_start.split(':')[0])
                m_end = int(setting.morning_end.split(':')[0])
                e_start = int(setting.evening_start.split(':')[0])
                e_end = int(setting.evening_end.split(':')[0])
            else:
                m_start = int(DEFAULT_SHIFT_SETTINGS["morning_start"].split(':')[0])
                m_end = int(DEFAULT_SHIFT_SETTINGS["morning_end"].split(':')[0])
                e_start = int(DEFAULT_SHIFT_SETTINGS["evening_start"].split(':')[0])
                e_end = int(DEFAULT_SHIFT_SETTINGS["evening_end"].split(':')[0])

            if m_start <= current_hour < m_end:
                auto_open_shift(fallback_db, "Sang")
                logger.info("[HARD RESET SUCCESS]: Đã tự động mở lại Ca Sáng thành công!")
            elif e_start <= current_hour < e_end:
                auto_open_shift(fallback_db, "Toi")
                logger.info("[HARD RESET SUCCESS]: Đã tự động mở lại Ca Tối thành công!")
            else:
                logger.info("[HARD RESET INFO]: Thời điểm reset nằm ngoài giờ trực hành chính.")

        except Exception as fallback_err:
            fallback_db.rollback()
            logger.error(f"[HARD RESET ORM ERROR]: Không thể seed dữ liệu hoặc mở ca trực: {str(fallback_err)}")
        finally:
            fallback_db.close()

        logger.warning(f"[HARD RESET SUCCESS]: Hệ thống đã được đưa về trạng thái nguyên bản sạch sẽ hoàn toàn bởi User {current_user.username}.")
        return {
            "status": "Success",
            "message": "Đã xóa sạch toàn bộ cấu trúc cũ và thiết lập lại Database sạch sẽ 100%! Sử dụng tài khoản admin / 123456 để đăng nhập.",
            "emergency_safety_backup_url": safety_drive_link
        }

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"[HARD RESET FATAL_ERROR]: Quy trình xóa trắng hệ thống thất bại: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Thất bại trong quá trình Hard Reset hệ thống: {str(e)}"
        )
    finally:
        _restore_lock.release()


@router_backup.post("/refresh-current-shift", status_code=status.HTTP_200_OK)
def refresh_current_shift_by_settings(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_privileged_user)
):
    try:
        # 🔥 ÉP KIỂU MÚI GIỜ CHÍ MẠNG: Khởi tạo múi giờ Việt Nam UTC+7
        vietnam_tz = timezone(timedelta(hours=7))
        current_hour = datetime.now(vietnam_tz).hour # Lấy giờ chuẩn theo múi giờ VN (Sẽ ra đúng 15h)
        
        logger.warning(f"[SHIFT REFRESH]: Người dùng {current_user.username} gọi lệnh làm mới ca trực tại mốc {current_hour}h.")

        setting = db.query(models.ShiftSetting).first()
        if setting:
            m_start = int(setting.morning_start.split(':')[0])
            m_end = int(setting.morning_end.split(':')[0])
            e_start = int(setting.evening_start.split(':')[0])
            e_end = int(setting.evening_end.split(':')[0])
        else:
            m_start = int(DEFAULT_SHIFT_SETTINGS["morning_start"].split(':')[0])
            m_end = int(DEFAULT_SHIFT_SETTINGS["morning_end"].split(':')[0])
            e_start = int(DEFAULT_SHIFT_SETTINGS["evening_start"].split(':')[0])
            e_end = int(DEFAULT_SHIFT_SETTINGS["evening_end"].split(':')[0])

        from services.shift_service import auto_open_shift
        activated_shift = None

        if m_start <= current_hour < m_end:
            activated_shift = "Sang"
            auto_open_shift(db, "Sang")
        elif e_start <= current_hour < e_end:
            activated_shift = "Toi"
            auto_open_shift(db, "Toi")

        if activated_shift:
            return {
                "status": "Success",
                "message": f"Đã đồng bộ khung giờ thành công! Hệ thống đã tự động bật phiên làm việc cho ca '{activated_shift}' dựa trên cấu hình mới.",
                "details": {"execution_time": f"{current_hour}h", "activated_shift": activated_shift}
            }
        else:
            return {
                "status": "Info",
                "message": "Đã chạy lệnh rà soát, giờ hiện tại nằm ngoài khung giờ ca trực.",
                "details": {"execution_time": f"{current_hour}h", "activated_shift": None}
            }
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e))