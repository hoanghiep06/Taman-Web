# assets.py
import io
import re
from typing import List, Optional, Tuple
from fastapi import APIRouter, Depends, HTTPException, status, UploadFile, File, Query
from sqlalchemy import func
from sqlalchemy.orm import Session, joinedload
from openpyxl import load_workbook

from database import get_db
from models import Asset, Elder, Room, Zone, Facility, User, Shift, InspectionLog, ElderHealthProfile
from schemas import AssetCreate, AssetResponse, RoleType, AssetStatsResponse, RoomPatrolProgressResponse
from core.dependencies import require_care_team, get_current_user

router = APIRouter(prefix="/api/admin/assets", tags=["5. [Quản lý] Danh Mục Tư Trang & Tài Sản"])

@router.get(
    "",
    response_model=List[AssetResponse],
    summary="Lấy danh sách Tư trang / Tài sản",
    description="""
    **Mô tả dành cho Frontend:**
    - Trả về danh sách toàn bộ tài sản trong hệ thống.
    - **Phân quyền Đa cơ sở:** Manager/NVCS thuộc Cơ sở nào thì mặc định chỉ thấy tài sản thuộc Cơ sở đó. Manager Vùng/Admin (`facility_id is None`) có thể xem toàn bộ hoặc lọc theo `facility_id`.
    - **Bộ lọc nâng cao:** 
      - `room_id`: Lọc tài sản trong 1 Phòng cụ thể.
      - `elder_id`: Lọc tất cả tư trang thuộc sở hữu của 1 Cụ.
      - `requires_inspection`: Trạng thái lọc đồ `True` (Cần NVCS đi tuần chụp ảnh) hoặc `False` (Đồ dùng lặt vặt).
    """
)
def get_all_assets(
    room_id: Optional[int] = Query(None, description="ID Phòng cần lọc"),
    elder_id: Optional[int] = Query(None, description="ID Cụ già cần xem danh mục tư trang riêng"),
    facility_id: Optional[int] = Query(None, description="ID Cơ sở (Dành cho Admin/Manager Vùng)"),
    requires_inspection: Optional[bool] = Query(None, description="Lọc theo yêu cầu đi tuần chụp ảnh (True/False)"),
    status_filter: Optional[str] = Query("Active", description="Trạng thái tài sản: 'Active' hoặc 'Archived'"),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):

    query = db.query(Asset).options(
        joinedload(Asset.elder),
        joinedload(Asset.room).joinedload(Room.zone).joinedload(Zone.facility)
    )

    # 1. Phân quyền đa cơ sở
    target_facility_id = current_user.facility_id if current_user.facility_id is not None else facility_id
    if target_facility_id is not None:
        query = query.join(Room).join(Zone).filter(Zone.facility_id == target_facility_id)

    # 2. Bộ lọc tùy chọn
    if room_id:
        query = query.filter(Asset.room_id == room_id)
    if elder_id:
        query = query.filter(Asset.elder_id == elder_id)
    if requires_inspection is not None:
        query = query.filter(Asset.requires_inspection == requires_inspection)
    if status_filter:
        query = query.filter(Asset.status == status_filter)

    assets = query.order_by(Asset.room_id, Asset.elder_id).all()

    # Đóng gói Response chứa đủ ngữ cảnh hiển thị
    results = []
    for a in assets:
        room_obj = a.room
        zone_obj = room_obj.zone if room_obj else None
        facility_obj = zone_obj.facility if zone_obj else None

        results.append(
            AssetResponse(
                id=a.id,
                asset_name=a.asset_name,
                room_id=a.room_id,
                elder_id=a.elder_id,
                contract_id=a.contract_id,
                requires_inspection=a.requires_inspection,
                status=a.status,
                created_at=a.created_at,
                elder_name=a.elder.full_name if a.elder else "Tài sản chung của phòng",
                room_number=room_obj.room_number if room_obj else "N/A",
                facility_name=facility_obj.name if facility_obj else "N/A"
            )
        )

    return results





# assets.py
from models import Shift, InspectionLog # Nhớ import thêm 2 model này ở đầu file nếu chưa có

@router.get(
    "/stats",
    response_model=AssetStatsResponse,
    summary="Thống kê Số lượng & Tiến độ Kiểm kê Tài sản (Cơ sở / Phân khu / Phòng)",
    description="""
    **API THÔNG MINH LẤY TỔNG SỐ LƯỢNG & TIẾN ĐỘ KIỂM KÊ:**
    - Cho phép truyền linh hoạt bộ lọc: `facility_id`, `zone_id`, `room_id`, `elder_id`.
    - Trả về chi tiết:
      + `total_assets`: Tổng số đồ active.
      + `total_required_inspection`: Tổng đồ BẮT BUỘC kiểm kê (cần chụp ảnh/báo mất).
      + `total_optional`: Tổng đồ dùng lặt vặt (không bắt buộc kiểm kê).
      + `inspected_required`: Số đồ bắt buộc ĐÃ KIỂM KÊ trong ca live hiện tại.
      + `inspected_total`: Tổng số đồ ĐÃ KIỂM KÊ trong ca live.
      + `required_percentage`: % tiến độ đi tuần ca trực (% nước ngập).
      + `is_completed`: `True` nếu đã kiểm kê xong 100% đồ bắt buộc.
    """
)
def get_asset_stats_and_progress(
    facility_id: Optional[int] = Query(None, description="Lọc theo ID Cơ sở"),
    zone_id: Optional[int] = Query(None, description="Lọc theo ID Phân khu (Khu A, B, C...)"),
    room_id: Optional[int] = Query(None, description="Lọc theo ID Phòng đích danh"),
    elder_id: Optional[int] = Query(None, description="Lọc theo ID Cụ già sở hữu"),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    # 1. Phân quyền đa cơ sở
    target_facility_id = current_user.facility_id if current_user.facility_id is not None else facility_id

    # 2. Base Query danh mục Tài sản Active
    asset_query = db.query(Asset).join(Room).join(Zone)

    if target_facility_id is not None:
        asset_query = asset_query.filter(Zone.facility_id == target_facility_id)
    if zone_id:
        asset_query = asset_query.filter(Room.zone_id == zone_id)
    if room_id:
        asset_query = asset_query.filter(Asset.room_id == room_id)
    if elder_id:
        asset_query = asset_query.filter(Asset.elder_id == elder_id)

    all_assets = asset_query.filter(Asset.status == "Active").all()
    all_asset_ids = [a.id for a in all_assets]

    # Phân loại tổng sản phẩm
    total_assets = len(all_assets)
    required_assets = [a for a in all_assets if a.requires_inspection]
    optional_assets = [a for a in all_assets if not a.requires_inspection]

    total_required = len(required_assets)
    total_optional = len(optional_assets)
    required_asset_ids = set(a.id for a in required_assets)

    # 3. Xác định Ca trực Live đang Mở (Shift.status == 'Open')
    active_shift = db.query(Shift).filter(Shift.status == "Open").order_by(Shift.id.desc()).first()

    inspected_required_count = 0
    inspected_total_count = 0
    shift_info = None

    if active_shift:
        shift_info = {
            "shift_id": active_shift.id,
            "shift_date": str(active_shift.shift_date),
            "shift_type": active_shift.shift_type
        }

        # Query các log kiểm kê mới nhất trong ca này có trạng thái hợp lệ ('Xanh' - Đã nộp, 'Vang' - Báo mất)
        logs = db.query(InspectionLog).filter(
            InspectionLog.shift_id == active_shift.id,
            InspectionLog.is_latest == True,
            InspectionLog.asset_id.in_(all_asset_ids),
            InspectionLog.status.in_(["Xanh", "Vang", "Success", "Missing"])
        ).all()

        inspected_asset_ids = set(l.asset_id for l in logs)
        inspected_total_count = len(inspected_asset_ids)
        
        # Đếm số đồ BẮT BUỘC kiểm kê đã hoàn thành
        inspected_required_count = len(inspected_asset_ids.intersection(required_asset_ids))

    # 4. Tính toán phần trăm tiến độ
    uninspected_required = max(0, total_required - inspected_required_count)
    
    if total_required > 0:
        required_pct = min(100.0, round((inspected_required_count / total_required) * 100, 1))
        is_completed = (inspected_required_count >= total_required)
    else:
        # Nếu phòng không có đồ bắt buộc kiểm kê -> Coi như hoàn thành 100%
        required_pct = 100.0
        is_completed = True

    total_pct = min(100.0, round((inspected_total_count / total_assets) * 100, 1)) if total_assets > 0 else 100.0

    return {
        "scope": {
            "facility_id": target_facility_id,
            "zone_id": zone_id,
            "room_id": room_id,
            "elder_id": elder_id
        },
        "active_shift": shift_info,
        "counts": {
            "total_assets": total_assets,                         # Tổng toàn bộ đồ
            "total_required_inspection": total_required,          # Tổng đồ BẮT BUỘC kiểm
            "total_optional_inspection": total_optional,          # Tổng đồ KHÔNG BẮT BUỘC kiểm
            "inspected_required": inspected_required_count,       # Đồ bắt buộc ĐÃ KIỂM KÊ
            "uninspected_required": uninspected_required,         # Đồ bắt buộc CHƯA KIỂM KÊ
            "inspected_total": inspected_total_count              # Tổng đồ đã kiểm kê (cả 2 loại)
        },
        "progress": {
            "required_percentage": required_pct,                  # % tiến độ đồ cần kiểm (Dùng cho hũ nước ngập UI)
            "total_percentage": total_pct,                        # % tiến độ toàn bộ đồ
            "is_completed": is_completed                          # True khi đã xong 100%
        }
    }




@router.get(
    "/rooms-progress",
    response_model=List[RoomPatrolProgressResponse],
    summary="Lấy danh sách Tất cả các Phòng kèm Tiến độ Kiểm kê (Cấp Cơ sở / Phân khu)",
    description="""
    **ENDPOINT TỔNG HỢP TIẾN ĐỘ ĐI TUẦN CHO TỪNG PHÒNG:**
    - Trả về danh sách từng Phòng trong Cơ sở.
    - Mỗi phòng đính kèm chi tiết: `total_required_inspection`, `inspected_count`, `% tiến độ nước ngập`, `is_completed`.
    - **Phân quyền Đa cơ sở:** Nhân viên/Manager thuộc Cơ sở nào sẽ tự động lấy danh sách phòng của Cơ sở đó. Admin/Manager Vùng có thể truyền `facility_id` hoặc `zone_id` tùy chọn.
    """
)
def get_rooms_patrol_progress(
    facility_id: Optional[int] = Query(None, description="ID Cơ sở (Dành cho Admin/Manager Vùng)"),
    zone_id: Optional[int] = Query(None, description="Lọc riêng theo Phân khu (Khu A, Khu B...)"),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    # 1. Phân quyền truy cập Cơ sở
    target_facility_id = current_user.facility_id if current_user.facility_id is not None else facility_id

    # 2. Query danh sách Phòng kèm Phân khu và Cơ sở
    room_query = db.query(Room).join(Zone, Room.zone_id == Zone.id).join(Facility, Zone.facility_id == Facility.id)

    if target_facility_id is not None:
        room_query = room_query.filter(Zone.facility_id == target_facility_id)
    if zone_id:
        room_query = room_query.filter(Room.zone_id == zone_id)

    rooms = room_query.order_by(Facility.id, Zone.name, Room.room_number).all()

    # 3. Lấy Ca trực Live đang Mở (Shift.status == 'Open')
    active_shift = db.query(Shift).filter(Shift.status == "Open").order_by(Shift.id.desc()).first()

    # 4. Tính toán số liệu từng phòng
    results = []

    for room in rooms:
        zone = room.zone
        facility = zone.facility if zone else None

        # Lấy toàn bộ tài sản Active trong phòng
        assets = db.query(Asset).filter(Asset.room_id == room.id, Asset.status == "Active").all()
        
        total_assets = len(assets)
        required_assets = [a for a in assets if a.requires_inspection]
        optional_assets = [a for a in assets if not a.requires_inspection]

        total_required = len(required_assets)
        total_optional = len(optional_assets)
        required_asset_ids = set(a.id for a in required_assets)

        inspected_count = 0

        # Nếu có ca trực live đang Mở -> Đếm số đồ bắt buộc đã được chụp ảnh/báo mất
        if active_shift and total_required > 0:
            logs = db.query(InspectionLog).filter(
                InspectionLog.shift_id == active_shift.id,
                InspectionLog.is_latest == True,
                InspectionLog.asset_id.in_(list(required_asset_ids)),
                InspectionLog.status.in_(["Xanh", "Vang", "Success", "Missing"])
            ).all()

            inspected_count = len(set(l.asset_id for l in logs))

        uninspected_count = max(0, total_required - inspected_count)

        # Tính phần trăm % nước ngập
        if total_required > 0:
            progress_pct = min(100.0, round((inspected_count / total_required) * 100, 1))
            is_completed = (inspected_count >= total_required)
        else:
            progress_pct = 100.0
            is_completed = True  # Phòng trống hoặc không có đồ bắt buộc kiểm kê -> Coi như Xong

        results.append(
            RoomPatrolProgressResponse(
                room_id=room.id,
                room_number=room.room_number,
                description=room.description,
                zone_id=zone.id if zone else 0,
                zone_name=zone.name if zone else "N/A",
                facility_id=facility.id if facility else 0,
                facility_name=facility.name if facility else "N/A",
                total_assets=total_assets,
                total_required_inspection=total_required,
                total_optional_inspection=total_optional,
                inspected_count=inspected_count,
                uninspected_count=uninspected_count,
                progress_percentage=progress_pct,
                is_completed=is_completed
            )
        )

    return results


# =========================================================================
# 2. CREATE: KHAI BÁO TƯ TRANG / TÀI SẢN MỚI
# =========================================================================
@router.post(
    "",
    response_model=AssetResponse,
    status_code=status.HTTP_201_CREATED,
    summary="Khai báo Tư trang / Tài sản mới",
    description="""
    **Mô tả dành cho Frontend:**
    - Sử dụng khi khai báo đồ đạc mới cho Cụ hoặc thêm thiết bị dùng chung cho Phòng.
    - **Tự động đồng bộ Phòng:** Nếu có chọn `elder_id` (tư trang riêng của Cụ), hệ thống sẽ **tự động lấy `room_id` theo đúng Phòng Cụ đó đang ở**, không lo nhân viên nhập lệch phòng[cite: 12].
    - `requires_inspection`: Tích chọn `True` nếu món đồ này giá trị cao/quan trọng (máy trợ thính, xe lăn, điện thoại...) cần NVCS chụp ảnh mỗi ca. Tích `False` nếu là đồ lặt vặt sinh hoạt.
    """
)
def create_asset(
    asset_in: AssetCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_care_team)
):
    asset_data = asset_in.model_dump()

    # Tự động đồng bộ room_id theo Cụ sở hữu
    if asset_in.elder_id:
        elder = db.query(Elder).options(joinedload(Elder.room)).filter(Elder.id == asset_in.elder_id).first()
        if not elder:
            raise HTTPException(status_code=404, detail="Không tìm thấy Cụ già sở hữu tài sản này!")
        
        if elder.room_id:
            asset_data["room_id"] = elder.room_id

    # Kiểm tra quyền cơ sở của phòng
    room = db.query(Room).options(joinedload(Room.zone)).filter(Room.id == asset_data["room_id"]).first()
    if not room:
        raise HTTPException(status_code=404, detail="Phòng lưu trữ tài sản không tồn tại!")

    if current_user.facility_id is not None and room.zone.facility_id != current_user.facility_id:
        raise HTTPException(status_code=403, detail="Bạn không có quyền thêm tài sản vào Cơ sở khác!")

    new_asset = Asset(**asset_data)
    db.add(new_asset)
    db.commit()
    db.refresh(new_asset)

    return AssetResponse(
        id=new_asset.id,
        asset_name=new_asset.asset_name,
        room_id=new_asset.room_id,
        elder_id=new_asset.elder_id,
        contract_id=new_asset.contract_id,
        requires_inspection=new_asset.requires_inspection,
        status=new_asset.status,
        created_at=new_asset.created_at,
        elder_name=new_asset.elder.full_name if new_asset.elder else "Tài sản chung của phòng",
        room_number=room.room_number,
        facility_name=room.zone.facility.name if (room.zone and room.zone.facility) else "N/A"
    )


# =========================================================================
# 3. READ DETAIL: XEM CHI TIẾT 1 MÓN TƯ TRANG
# =========================================================================
@router.get(
    "/{asset_id}",
    response_model=AssetResponse,
    summary="Xem chi tiết 1 món Tư trang / Tài sản",
    description="Lấy đầy đủ thuộc tính của 1 tài sản dựa vào ID (gồm Cụ sở hữu, Phòng, Cơ sở và Yêu cầu kiểm kê)."
)
def get_asset_by_id(
    asset_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    asset = db.query(Asset).options(
        joinedload(Asset.elder),
        joinedload(Asset.room).joinedload(Room.zone).joinedload(Zone.facility)
    ).filter(Asset.id == asset_id).first()

    if not asset:
        raise HTTPException(status_code=404, detail="Không tìm thấy tài sản này!")

    room_obj = asset.room
    zone_obj = room_obj.zone if room_obj else None

    return AssetResponse(
        id=asset.id,
        asset_name=asset.asset_name,
        room_id=asset.room_id,
        elder_id=asset.elder_id,
        contract_id=asset.contract_id,
        requires_inspection=asset.requires_inspection,
        status=asset.status,
        created_at=asset.created_at,
        elder_name=asset.elder.full_name if asset.elder else "Tài sản chung của phòng",
        room_number=room_obj.room_number if room_obj else "N/A",
        facility_name=zone_obj.facility.name if (zone_obj and zone_obj.facility) else "N/A"
    )


# =========================================================================
# 4. UPDATE: CẬP NHẬT TƯ TRANG / CHUYỂN CHỦ SỞ HỮU
# =========================================================================
@router.put(
    "/{asset_id}",
    response_model=AssetResponse,
    summary="Cập nhật thông tin Tư trang / Tài sản",
    description="""
    **Mô tả dành cho Frontend:**
    - Cho phép đổi tên món đồ, chuyển quyền sở hữu cho Cụ khác, hoặc thay đổi cờ `requires_inspection` (Bật/Tắt chế độ NVCS đi tuần chụp ảnh).
    - Tự động di dời `room_id` của tài sản theo đúng phòng của Cụ sở hữu mới.
    """
)
def update_asset(
    asset_id: int,
    asset_data: AssetCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_care_team)
):
    asset = db.query(Asset).filter(Asset.id == asset_id).first()
    if not asset:
        raise HTTPException(status_code=404, detail="Không tìm thấy tài sản để cập nhật!")

    # Tự động cập nhật phòng theo Cụ sở hữu mới
    if asset_data.elder_id:
        elder = db.query(Elder).filter(Elder.id == asset_data.elder_id).first()
        if elder and elder.room_id:
            asset_data.room_id = elder.room_id

    for field, value in asset_data.model_dump(exclude_unset=True).items():
        setattr(asset, field, value)

    db.commit()
    db.refresh(asset)

    room_obj = db.query(Room).options(joinedload(Room.zone).joinedload(Zone.facility)).filter(Room.id == asset.room_id).first()

    return AssetResponse(
        id=asset.id,
        asset_name=asset.asset_name,
        room_id=asset.room_id,
        elder_id=asset.elder_id,
        contract_id=asset.contract_id,
        requires_inspection=asset.requires_inspection,
        status=asset.status,
        created_at=asset.created_at,
        elder_name=asset.elder.full_name if asset.elder else "Tài sản chung của phòng",
        room_number=room_obj.room_number if room_obj else "N/A",
        facility_name=room_obj.zone.facility.name if (room_obj and room_obj.zone and room_obj.zone.facility) else "N/A"
    )


# =========================================================================
# 5. DELETE: XÓA TÀI SẢN (CHUYỂN VỀ ARCHIVED HOẶC XÓA CỨNG)
# =========================================================================
@router.delete(
    "/{asset_id}",
    status_code=status.HTTP_200_OK,
    summary="Xóa hoặc Chuyển kho lưu trữ (Archive) tài sản",
    description="Xóa món tư trang khỏi danh mục quản lý active của phòng."
)
def delete_asset(
    asset_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_care_team)
):
    asset = db.query(Asset).filter(Asset.id == asset_id).first()
    if not asset:
        raise HTTPException(status_code=404, detail="Không tìm thấy tài sản để xóa!")

    db.delete(asset)
    db.commit()
    return {"message": f"Đã xóa thành công tài sản '{asset.asset_name}'"}



# =========================================================================
# HELPER: BÓC TÁCH PHÂN KHU VÀ SỐ PHÒNG TỰ ĐỘNG
# =========================================================================
import re
from typing import Tuple

def parse_zone_and_room(room_str: str) -> Tuple[str, str]:
    """
    Bóc tách tên phân khu và số phòng từ chuỗi nhập liệu.
    """
    cleaned = room_str.strip()
    
    # Regex tách phân khu (Group 1) và số phòng (Group 2)
    match = re.match(r"^(?:Khu\s*)?([A-Za-z]+)[\s\-_]*(.*)$", cleaned, re.IGNORECASE)
    
    if match:
        zone_name = match.group(1).upper()
        room_name = match.group(2).strip()
        
        # Phòng hờ trường hợp chuỗi nhập chỉ có chữ, không có số phòng (VD: "Khu A")
        if not room_name:
            room_name = cleaned
            
        return zone_name, room_name
        
    return "Chung", cleaned


# =========================================================================
# HELPER: TỰ ĐỘNG ĐẢM BẢO HỒ SƠ SỨC KHỎE CHO CỤ
# =========================================================================
def ensure_elder_health_profile(db: Session, elder_id: int):
    profile = db.query(ElderHealthProfile).filter(ElderHealthProfile.elder_id == elder_id).first()
    if not profile:
        profile = ElderHealthProfile(
            elder_id=elder_id,
            has_surgery=False,
            has_fall=False,
            has_stroke=False,
            has_cardiovascular=False,
            drug_allergies=[],
            food_allergies=[],
            chronic_diseases=[]
        )
        db.add(profile)
        db.flush()


# =========================================================================
# ENDPOINT: IMPORT MA TRẬN TƯ TRANG / VẬT TƯ KIỂM KÊ TỪ EXCEL
# =========================================================================
@router.post(
    "/import-xlsx",
    status_code=status.HTTP_200_OK,
    summary="Import danh mục Tư trang ma trận từ file Excel",
    description="""
    **Quy chuẩn cấu trúc file Excel Ma trận:**
    - Cột 0 (A): Tên Cơ sở (Hỗ trợ merge hàng)
    - Cột 1 (B): Tên/Số Phòng (VD: 'B13' -> tự tách Khu B, Phòng 13; hỗ trợ merge hàng)
    - Cột 2 (C): Họ và tên Cụ
    - Cột 3 trở đi: Tên các món tài sản / vật tư kiểm kê (đánh dấu x, v, 1, true để kích hoạt)
    """
)
async def import_assets_from_xlsx(
    file: UploadFile = File(..., description="File Excel ma trận tư trang (.xlsx/.xls)"),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_care_team)
):
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST, 
            detail="Hệ thống chỉ chấp nhận file Excel định dạng .xlsx hoặc .xls!"
        )

    try:
        contents = await file.read()
        wb = load_workbook(io.BytesIO(contents), data_only=True)
        ws = wb.active

        # Đọc dòng tiêu đề (Header)
        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
        headers = [str(cell).strip() if cell is not None else "" for cell in header_row]

        processed_elders = 0
        total_assets_upserted = 0

        # Biến lưu trữ ngữ cảnh khi duyệt qua các ô bị merge
        current_facility_name: Optional[str] = None
        current_room_raw: Optional[str] = None

        # Danh sách các tên cột định danh cần bỏ qua khi duyệt tài sản
        ignored_header_keywords = [
            "cơ sở", "co so", "facility",
            "phòng", "phong", "room",
            "tên cụ", "ten cu", "họ và tên", "ho va ten", "elder", "stt"
        ]

        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or all(cell is None for cell in row):
                continue

            facility_cell = row[0] if len(row) > 0 else None
            room_cell = row[1] if len(row) > 1 else None
            elder_cell = row[2] if len(row) > 2 else None

            # Cập nhật context từ các ô merge
            if facility_cell is not None and str(facility_cell).strip() != "":
                current_facility_name = str(facility_cell).strip()

            if room_cell is not None and str(room_cell).strip() != "":
                current_room_raw = str(room_cell).strip()

            # Bỏ qua nếu dòng không có tên Cụ
            if elder_cell is None or str(elder_cell).strip() == "":
                continue

            elder_name = str(elder_cell).strip()

            # -------------------------------------------------------------
            # BƯỚC 1: XỬ LÝ CƠ SỞ (FACILITY)
            # -------------------------------------------------------------
            target_facility = None
            if current_facility_name:
                target_facility = db.query(Facility).filter(
                    func.lower(Facility.name) == current_facility_name.lower()
                ).first()

                if not target_facility:
                    target_facility = Facility(
                        name=current_facility_name,
                        address="Khởi tạo tự động từ Import Excel"
                    )
                    db.add(target_facility)
                    db.flush()
            elif current_user.facility_id:
                target_facility = db.query(Facility).filter(Facility.id == current_user.facility_id).first()
            else:
                target_facility = db.query(Facility).first()
                if not target_facility:
                    target_facility = Facility(name="Cơ sở Mặc định", address="Chưa cập nhật")
                    db.add(target_facility)
                    db.flush()

            # Phân quyền cơ sở đối với Manager quản lý đơn cơ sở
            if current_user.facility_id is not None and target_facility.id != current_user.facility_id:
                raise HTTPException(
                    status_code=status.HTTP_403_FORBIDDEN,
                    detail=f"Bạn không có quyền import dữ liệu vào Cơ sở '{target_facility.name}'!"
                )

            # -------------------------------------------------------------
            # BƯỚC 2: XỬ LÝ PHÂN KHU (ZONE) & PHÒNG (ROOM)
            # -------------------------------------------------------------
            room_str = current_room_raw if current_room_raw else "Phòng Chung"
            zone_name, room_number = parse_zone_and_room(room_str)

            # 2.1. Đảm bảo Phân khu tồn tại trong Cơ sở
            zone = db.query(Zone).filter(
                Zone.facility_id == target_facility.id,
                func.lower(Zone.name) == zone_name.lower()
            ).first()

            if not zone:
                zone = Zone(
                    facility_id=target_facility.id,
                    name=zone_name,
                    description=f"{zone_name} tạo tự động từ Import"
                )
                db.add(zone)
                db.flush()

            # 2.2. Đảm bảo Phòng tồn tại trong Phân khu
            room = db.query(Room).filter(
                Room.zone_id == zone.id,
                func.lower(Room.room_number) == room_number.lower()
            ).first()

            if not room:
                room = Room(
                    zone_id=zone.id,
                    room_number=room_number,
                    description=f"Phòng {room_number} ({zone.name} - {target_facility.name})"
                )
                db.add(room)
                db.flush()

            # -------------------------------------------------------------
            # BƯỚC 3: XỬ LÝ HỒ SƠ CỤ & BẢNG THEO DÕI SỨC KHỎE
            # -------------------------------------------------------------
            elder = db.query(Elder).filter(
                func.lower(Elder.full_name) == elder_name.lower(),
                Elder.room_id == room.id
            ).first()

            if not elder:
                elder = Elder(
                    full_name=elder_name,
                    room_id=room.id
                )
                db.add(elder)
                db.flush()

            # Tự động khởi tạo Hồ sơ sức khỏe nếu chưa có
            ensure_elder_health_profile(db, elder.id)
            processed_elders += 1

            # -------------------------------------------------------------
            # BƯỚC 4: XỬ LÝ MA TRẬN VẬT TƯ / TÀI SẢN (TỪ CỘT 3 TRỞ ĐI)
            # -------------------------------------------------------------
            existing_assets = db.query(Asset).filter(Asset.elder_id == elder.id).all()
            asset_dict = {a.asset_name.lower().strip(): a for a in existing_assets}

            for col_index in range(3, len(row)):
                if col_index >= len(headers):
                    break

                asset_name = headers[col_index]
                if not asset_name:
                    continue

                # Bỏ qua các cột tiêu đề trùng tên định danh
                if any(kw in asset_name.lower() for kw in ignored_header_keywords):
                    continue

                cell_val = row[col_index]
                is_checked = (cell_val is True) or (
                    str(cell_val).strip().upper() in ["TRUE", "1", "1.0", "X", "V", "✓", "YES", "CÓ"]
                )

                if is_checked:
                    search_key = asset_name.lower().strip()
                    if search_key in asset_dict:
                        target_asset = asset_dict[search_key]
                        target_asset.status = "Active"
                        target_asset.room_id = room.id
                        target_asset.requires_inspection = True
                    else:
                        new_asset = Asset(
                            asset_name=asset_name.strip(),
                            room_id=room.id,
                            elder_id=elder.id,
                            status="Active",
                            requires_inspection=True
                        )
                        db.add(new_asset)
                    total_assets_upserted += 1

        db.commit()

        return {
            "status": "Success",
            "message": "Đồng bộ ma trận cơ sở, phòng và vật tư kiểm kê từ Excel thành công!",
            "summary": {
                "total_elders_processed": processed_elders,
                "total_assets_active": total_assets_upserted
            }
        }

    except HTTPException:
        db.rollback()
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Lỗi cấu trúc hoặc xử lý dữ liệu file Excel: {str(e)}"
        )

    
# =========================================================================
# 7. LẤY TOÀN BỘ TƯ TRANG / TÀI SẢN THEO PHÒNG ĐÍCH DANH
# =========================================================================
@router.get(
    "/room/{room_id}",
    response_model=List[AssetResponse],
    summary="Lấy danh sách Tư trang / Tài sản theo ID Phòng",
    description="""
    **Mô tả dành cho Frontend:**
    - Trả về tất cả các món tư trang (cả đồ cá nhân của các Cụ và đồ dùng chung) nằm trong `room_id` này.
    - Dùng cho màn hình Quản lý Phòng để hiển thị danh mục thiết bị/tư trang đang có trong phòng.
    """
)
def get_assets_by_room_id(
    room_id: int,
    requires_inspection_only: bool = Query(False, description="Nếu True: Chỉ lấy đồ BẮT BUỘC kiểm kê đi tuần"),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    room = db.query(Room).filter(Room.id == room_id).first()
    if not room:
        raise HTTPException(status_code=404, detail="Không tìm thấy phòng này!")

    query = db.query(Asset).options(
        joinedload(Asset.elder),
        joinedload(Asset.room).joinedload(Room.zone).joinedload(Zone.facility)
    ).filter(Asset.room_id == room_id, Asset.status == "Active")

    if requires_inspection_only:
        query = query.filter(Asset.requires_inspection == True)

    assets = query.order_by(Asset.elder_id, Asset.asset_name).all()

    results = []
    for a in assets:
        room_obj = a.room
        zone_obj = room_obj.zone if room_obj else None
        facility_obj = zone_obj.facility if zone_obj else None

        results.append(
            AssetResponse(
                id=a.id,
                asset_name=a.asset_name,
                room_id=a.room_id,
                elder_id=a.elder_id,
                contract_id=a.contract_id,
                requires_inspection=a.requires_inspection,
                status=a.status,
                created_at=a.created_at,
                elder_name=a.elder.full_name if a.elder else "Tài sản chung của phòng",
                room_number=room_obj.room_number if room_obj else "N/A",
                facility_name=facility_obj.name if facility_obj else "N/A"
            )
        )
    return results


# =========================================================================
# LẤY TOÀN BỘ TƯ TRANG / TÀI SẢN THEO CƠ SỞ ĐÍCH DANH
# =========================================================================
@router.get(
    "/facility/{facility_id}",
    response_model=List[AssetResponse],
    summary="Lấy danh sách Tư trang / Tài sản theo ID Cơ sở",
    description="""
    **Mô tả dành cho Frontend:**
    - Trả về toàn bộ tư trang thuộc về tất cả các Phòng/Khu nằm trong `facility_id` được chọn.
    - Dùng cho Manager/Admin xem Báo cáo tổng kiểm kê tài sản cấp Cơ sở.
    """
)
def get_assets_by_facility_id(
    facility_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    # Kiểm tra quyền truy cập cơ sở
    if current_user.facility_id is not None and current_user.facility_id != facility_id:
        raise HTTPException(status_code=403, detail="Bạn không có quyền xem danh mục tài sản của Cơ sở khác!")

    facility = db.query(Facility).filter(Facility.id == facility_id).first()
    if not facility:
        raise HTTPException(status_code=404, detail="Không tìm thấy Cơ sở này!")

    assets = db.query(Asset).options(
        joinedload(Asset.elder),
        joinedload(Asset.room).joinedload(Room.zone).joinedload(Zone.facility)
    ).join(Room).join(Zone).filter(
        Zone.facility_id == facility_id,
        Asset.status == "Active"
    ).order_by(Asset.room_id, Asset.elder_id).all()

    results = []
    for a in assets:
        room_obj = a.room
        zone_obj = room_obj.zone if room_obj else None

        results.append(
            AssetResponse(
                id=a.id,
                asset_name=a.asset_name,
                room_id=a.room_id,
                elder_id=a.elder_id,
                contract_id=a.contract_id,
                requires_inspection=a.requires_inspection,
                status=a.status,
                created_at=a.created_at,
                elder_name=a.elder.full_name if a.elder else "Tài sản chung của phòng",
                room_number=room_obj.room_number if room_obj else "N/A",
                facility_name=facility.name
            )
        )
    return results