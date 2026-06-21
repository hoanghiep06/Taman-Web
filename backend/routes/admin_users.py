# routes/admin_users.py
import io 
import pytz
from zoneinfo import ZoneInfo
from fastapi import UploadFile, File, APIRouter, Depends, HTTPException, status
from sqlalchemy.orm import Session
from database import get_db

from openpyxl import load_workbook
from models import User, LoginLog, InspectionLog, AuditLog, Asset, Room, Shift
from schemas import UserCreate, UserResponse, RoleType
from core.dependencies import get_privileged_user  # ĐÃ SỬA: Mở rộng cho cả Admin và Manager cùng vào
from passlib.context import CryptContext
from typing import List


pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")

router_users = APIRouter(
    prefix="/admin/users",
    tags=["Quản lý Tài khoản (Admin/Manager)"],
    dependencies=[Depends(get_privileged_user)] # ĐÃ SỬA: Cho cả 2 cấp quyền tiếp cận menu này
)

# =========================================================================
# 1. API TẠO TÀI KHOẢN (ĐỒNG BỘ PHÂN CẤP BẬC)
# =========================================================================
@router_users.post("", response_model=UserResponse, status_code=status.HTTP_201_CREATED)
def create_user(
    user_in: UserCreate, 
    db: Session = Depends(get_db), 
    current_user: User = Depends(get_privileged_user)
):
    # 1. Kiểm tra trùng lặp tên đăng nhập
    if db.query(User).filter(User.username == user_in.username).first():
        raise HTTPException(status_code=400, detail="Tên đăng nhập đã tồn tại")

    final_role = user_in.role.value # Lấy chuỗi string ("Admin", "Manager", "Staff")
    
    # ──── HÀNG RÀO NHÂN SỰ 1: Chặn không cho Manager tạo ra cấp trên (Admin) ────
    if current_user.role == "Manager" and final_role == "Admin":
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Tài khoản cấp Quản lý (Manager) không có quyền hạn khởi tạo tài khoản cấp Quản trị viên (Admin)."
        )
    
    # Logic giữ nguyên: Chỉ tài khoản admin gốc tối cao mới được cấp Admin cho người khác (nếu cần kiểm soát chặt)
    if final_role == "Admin" and current_user.username != "admin":
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN, 
            detail="Chỉ tài khoản admin gốc của hệ thống mới được cấp quyền Quản trị viên cho người khác."
        )

    # 3. Băm mật khẩu và lưu
    hashed_password = pwd_context.hash(user_in.password)
    new_user = User(
        username=user_in.username,
        password_hash=hashed_password,
        full_name=user_in.full_name,
        role=final_role,
        is_active=user_in.is_active
    )
    db.add(new_user)
    db.commit()
    db.refresh(new_user)
    return new_user


# =========================================================================
# 2. API XEM DANH SÁCH (CẢ 2 ĐỀU XEM ĐƯỢC TOÀN BỘ)
# =========================================================================
@router_users.get("", response_model=List[UserResponse])
def get_all_users(db: Session = Depends(get_db), current_user: User = Depends(get_privileged_user)):
    return db.query(User).all()


# =========================================================================
# 3. API KHÓA/MỞ KHÓA TÀI KHOẢN (CHỐNG SỬA CHỨC TRÊN)
# =========================================================================
@router_users.put("/{user_id}/toggle-lock", response_model=UserResponse)
def toggle_lock_user(
    user_id: int, 
    db: Session = Depends(get_db),
    current_user: User = Depends(get_privileged_user)
):
    user = db.query(User).filter(User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="Không tìm thấy người dùng")
        
    if user.username == "admin":
        raise HTTPException(status_code=403, detail="Không thể khóa tài khoản biệt lập admin gốc")

    # ──── HÀNG RÀO NHÂN SỰ 2: Chặn không cho Manager tương tác khóa tài khoản của Admin ────
    if current_user.role == "Manager" and user.role == "Admin":
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Quy trình bị từ chối. Bạn không có quyền hạn thay đổi trạng thái hoạt động của Quản trị viên."
        )

    # Đảo ngược trạng thái hoạt động
    user.is_active = not user.is_active
    db.commit()
    db.refresh(user)
    return user


# =========================================================================
# 4. API XÓA TÀI KHOẢN (MANAGER KHÔNG XÓA ĐƯỢC ADMIN - ADMIN XÓA ĐƯỢC NHAU)
# =========================================================================
@router_users.delete("/{user_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_user(
    user_id: int, 
    db: Session = Depends(get_db),
    current_user: User = Depends(get_privileged_user)
):
    user = db.query(User).filter(User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="Không tìm thấy người dùng")
        
    if user.username == "admin":
        raise HTTPException(status_code=403, detail="Không thể xóa tài khoản biệt lập admin gốc")

    # ──── HÀNG RÀO NHÂN SỰ 3: Manager KHÔNG thể xóa Admin ────
    if current_user.role == "Manager" and user.role == "Admin":
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Cấp bậc vận hành không đủ thẩm quyền. Bạn không được phép xóa tài khoản của Quản trị viên."
        )
        
    # LƯU Ý ĐÚNG Ý BẠN: Nếu current_user.role == "Admin" và target user.role == "Admin", 
    # câu điều kiện if trên sẽ tự động bỏ qua, cho phép các Admin xóa lẫn nhau một cách hợp lệ.

    db.delete(user)
    db.commit()


# =========================================================================
# API: IMPORT DANH SÁCH NHÂN VIÊN TỰ ĐỘNG TỪ FILE EXCEL (UPSERT LOGIC)
# =========================================================================
@router_users.post("/import-xlsx", status_code=status.HTTP_200_OK)
async def import_staff_from_xlsx(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_privileged_user)
):
    """
    API Import danh sách tài khoản nhân viên hàng loạt từ file .xlsx:
    - Cột A (Index 0): STT (Hệ thống tự động bỏ qua)
    - Cột B (Index 1): Họ và tên nhân viên (Dùng làm full_name)
    - Cột C (Index 2): Số điện thoại (Dùng làm Username gốc & Mật khẩu khởi tạo)
    - ĐBỌC GIÁP: Tự động xử lý dấu nháy đơn `'` và tự động bù số 0 nếu bị mất.
    """
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST, 
            detail="Hệ thống chỉ chấp nhận file Excel định dạng .xlsx hoặc .xls"
        )

    try:
        file_contents = await file.read()
        wb = load_workbook(io.BytesIO(file_contents), data_only=True)
        ws = wb.active
        
        inserted_count = 0
        updated_count = 0

        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or len(row) < 3:
                continue

            name_raw = row[1]   # Cột B: Họ tên nhân viên
            phone_raw = row[2]  # Cột C: Số điện thoại

            if name_raw is None or phone_raw is None:
                continue

            full_name = str(name_raw).strip()
            
            # ──── BỘ LỌC CHUẨN HÓA SỐ ĐIỆN THOẠI THÔNG MINH (SANIDITY CHECK) ────
            phone_str = str(phone_raw).strip()
            
            # 1. Lột bỏ dấu nháy đơn nếu nó lọt trực tiếp vào chuỗi ký tự thô
            phone_str = phone_str.lstrip("'")
            
            # 2. Xử lý trường hợp quên gõ dấu ': Excel tự chuyển thành số (ví dụ: 901234567) -> Tự bù lại số 0
            if phone_str.isdigit() and len(phone_str) == 9:
                phone_str = "0" + phone_str
            
            username_sdt = phone_str

            if not full_name or not username_sdt:
                continue

            # 3. Tra cứu kiểm tra sự tồn tại của tài khoản dưới Database
            existing_user = db.query(User).filter(User.username == username_sdt).first()

            if existing_user:
                # Nếu nhân viên đã có tài khoản -> Chỉ cập nhật thông tin Họ tên mới
                existing_user.full_name = full_name
                updated_count += 1
            else:
                # Nếu nhân viên mới tinh -> Khởi tạo tài khoản bọc giáp mật mã
                hashed_password = pwd_context.hash(username_sdt)
                
                new_user = User(
                    username=username_sdt,
                    password_hash=hashed_password,
                    full_name=full_name,
                    role="Staff", 
                    is_active=True
                )
                db.add(new_user)
                inserted_count += 1

        db.commit()

        return {
            "status": "Success",
            "message": "Đồng bộ danh sách tài khoản nhân viên từ file Excel thành công chỉnh chu!",
            "summary": {
                "total_new_staff_created": inserted_count,
                "total_existing_staff_updated": updated_count
            }
        }

    except Exception as e:
        db.rollback()
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, 
            detail=f"Lỗi hệ thống trong quá trình bóc tách danh mục nhân sự: {str(e)}"
        )
    

# =========================================================================
# API: XEM HỒ SƠ LỊCH SỬ TỔNG HỢP CỦA 1 NHÂN VIÊN (CHỈ ADMIN/MANAGER)
# =========================================================================
@router_users.get("/{user_id}/comprehensive-history", status_code=status.HTTP_200_OK)
def get_user_comprehensive_history(
    user_id: int,
    limit: int = 50, # Giới hạn số lượng bản ghi gần đây để tránh kéo quá nhiều làm nặng RAM
    db: Session = Depends(get_db),
    current_user: User = Depends(get_privileged_user)
):
    """
    Endpoint đặc quyền dành cho Admin và Manager để tra cứu toàn bộ vết hoạt động của 1 nhân viên:
    1. Thông tin tài khoản tĩnh hiện tại
    2. Nhật ký đăng nhập gần đây (Thời gian, IP, Thiết bị)
    3. Nhật ký đi tuần/kiểm kê thực tế (Tên đồ vật, Phòng, Số ca, Trạng thái màu sắc)
    4. Nhật ký hành động hệ thống (Up ảnh, Xin mã bảo mật, Báo mất...)
    """
    # 1. Kiểm tra sự tồn tại của nhân viên mục tiêu
    user = db.query(User).filter(User.id == user_id).first()
    if not user:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND, 
            detail="Không tìm thấy thông tin nhân viên này trên hệ thống"
        )

    # Khởi tạo múi giờ Việt Nam để quy đổi mốc thời gian hiển thị sạch sẽ
    tz = ZoneInfo("Asia/Ho_Chi_Minh")

    # 2. TRÍCH XUẤT PHÂN HỆ 1: Lịch sử đăng nhập (Login Logs)
    login_logs = db.query(LoginLog).filter(LoginLog.user_id == user_id)\
        .order_by(LoginLog.login_time.desc()).limit(limit).all()

    login_history = [
        {
            "id": log.id,
            "login_at": log.login_time.astimezone(tz).strftime("%Y-%m-%d %H:%M:%S") if log.login_time else None,
            "ip_address": log.ip_address,
            "user_agent": log.user_agent # Thông tin thiết bị điện thoại/trình duyệt
        } for log in login_logs
    ]

    # 3. TRÍCH XUẤT PHÂN HỆ 2: Nhật ký nghiệp vụ đi tuần / nộp ảnh / báo mất
    # Thực hiện Explicit Join 4 bảng để lấy ra biên bản đầy đủ thông tin
    inspection_logs = db.query(
        InspectionLog, 
        Asset.asset_name, 
        Room.room_number, 
        Shift.shift_date, 
        Shift.shift_type
    ).join(Asset, InspectionLog.asset_id == Asset.id)\
     .join(Room, Asset.room_id == Room.id)\
     .join(Shift, InspectionLog.shift_id == Shift.id)\
     .filter(InspectionLog.user_id == user_id)\
     .order_by(InspectionLog.created_at.desc()).limit(limit).all()

    inspection_history = [
        {
            "log_id": log.id,
            "shift_date": str(s_date),
            "shift_type": s_type, # Ca Sang hoặc Ca Toi
            "room_number": r_num,
            "asset_name": asset_name,
            "status": log.status, # Xanh / Vang / Dang_Xu_Ly / Loi_Upload
            "note": log.note, # Lý do giải trình nếu là báo mất màu Vàng
            "version": log.version,
            "inspected_at": log.created_at.astimezone(tz).strftime("%Y-%m-%d %H:%M:%S") if log.created_at else None
        } for log, asset_name, r_num, s_date, s_type in inspection_logs
    ]

    # 4. TRÍCH XUẤT PHÂN HỆ 3: Nhật ký hành động lõi kiểm toán (Audit Logs)
    # Theo dõi vết bấm nút: REQUEST_NONCE, CHECKIN_REQUEST_ACCEPTED,...
    audit_logs = db.query(AuditLog).filter(AuditLog.actor_id == user_id)\
        .order_by(AuditLog.created_at.desc()).limit(limit).all()

    audit_history = [
        {
            "id": log.id,
            "action": log.action, # Tên hành động cụ thể
            "target_id": log.target_id,
            "ip_address": log.ip_address,
            "payload": log.payload, # Chi tiết dữ liệu đính kèm lúc hành động
            "created_at": log.created_at.astimezone(tz).strftime("%Y-%m-%d %H:%M:%S") if log.created_at else None
        } for log in audit_logs
    ]

    # 5. ĐÓNG GÓI BẢN TIN TRẢ VỀ TOÀN DIỆN
    return {
        "status": "Success",
        "staff_info": {
            "user_id": user.id,
            "username_sdt": user.username, # Số điện thoại đăng nhập
            "full_name": user.full_name,
            "role": user.role,
            "is_active": user.is_active
        },
        "statistics": {
            "loaded_limit": limit,
            "recent_logins_count": len(login_history),
            "recent_inspections_count": len(inspection_history),
            "recent_actions_count": len(audit_history)
        },
        "login_history": login_history,       # Cung cấp cho Tab 1 trên Frontend
        "inspection_history": inspection_history, # Cung cấp cho Tab 2 trên Frontend
        "audit_history": audit_history         # Cung cấp cho Tab 3 trên Frontend
    }