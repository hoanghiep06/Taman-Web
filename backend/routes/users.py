import io
import re
from zoneinfo import ZoneInfo
from typing import List, Optional, Dict
from fastapi import APIRouter, Depends, HTTPException, status, UploadFile, File, Query
from sqlalchemy.orm import Session, joinedload
from sqlalchemy import or_
from openpyxl import load_workbook

from database import get_db
from models import User, LoginLog, InspectionLog, AuditLog, Asset, Room, Shift, Facility
from schemas import (
    UserCreate, UserResponse, RoleType, UserUpdate, 
    BulkActionRequest, BulkLockRequest, UserResetPassword
)
from core.security import get_password_hash
from core.dependencies import get_current_user, require_management

router = APIRouter(prefix="/api/admin/users", tags=["2. [Admin/Manager] Quản lý Tài Khoản"])

ROOT_ADMIN_ID = 1  # ID của tài khoản Admin Gốc (Super Admin) cao nhất

def is_root_admin(user: User) -> bool:
    """Kiểm tra xem user hiện tại có phải là Tài khoản Quản trị viên Gốc hay không"""
    return user.role == RoleType.Admin and user.id == ROOT_ADMIN_ID


# =========================================================================
# 1. DANH SÁCH TÀI KHOẢN (SEARCH, FILTER & ẨN ADMIN ĐỐI VỚI MANAGER)
# =========================================================================
@router.get("", response_model=List[UserResponse])
def get_all_users(
    search: Optional[str] = Query(None, description="Tìm theo họ tên, username hoặc số điện thoại"),
    role: Optional[RoleType] = Query(None, description="Lọc theo vai trò (Role)"),
    facility_id: Optional[int] = Query(None, description="Lọc theo cơ sở làm việc"),
    is_active: Optional[bool] = Query(None, description="Lọc theo trạng thái Hoạt động / Đã khóa"),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_management)
):
    """
    Lấy danh sách người dùng hỗ trợ phân quyền & UI:
    - Manager: Tự động ẩn hoàn toàn tài khoản role Admin.
    - Hỗ trợ Search nhanh để gắn trực tiếp vào ô tìm kiếm trên Header bảng.
    - Trả về kèm facility_name để frontend render Badge cơ sở mượt mà.
    """
    query = db.query(User).options(joinedload(User.facility))

    # 1. Manager tuyệt đối không xem được danh sách tài khoản Admin
    if current_user.role == RoleType.Manager:
        query = query.filter(User.role != RoleType.Admin)

    # 2. Tìm kiếm đa năng (Tên, SĐT, Username)
    if search:
        search_pattern = f"%{search.strip()}%"
        query = query.filter(
            or_(
                User.full_name.ilike(search_pattern),
                User.username.ilike(search_pattern),
                User.phone_number.ilike(search_pattern)
            )
        )

    # 3. Bộ lọc nâng cao
    if role:
        query = query.filter(User.role == role)
    if facility_id is not None:
        query = query.filter(User.facility_id == facility_id)
    if is_active is not None:
        query = query.filter(User.is_active == is_active)

    users = query.order_by(User.id.desc()).all()

    # Format dữ liệu đầu ra
    results = []
    for u in users:
        fac_name = u.facility.name if (hasattr(u, 'facility') and u.facility) else None
        results.append(
            UserResponse(
                id=u.id,
                username=u.username,
                full_name=u.full_name,
                phone_number=u.phone_number,
                role=u.role,
                facility_id=u.facility_id,
                facility_name=fac_name,
                is_active=u.is_active,
                must_change_password=getattr(u, 'must_change_password', False),
                created_at=u.created_at
            )
        )

    return results


# =========================================================================
# 2. TẠO TÀI KHOẢN MỚI
# =========================================================================
@router.post("", response_model=UserResponse, status_code=status.HTTP_201_CREATED)
def create_user(
    user_in: UserCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_management)
):
    if db.query(User).filter(User.username == user_in.username).first():
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Tên đăng nhập đã tồn tại")

    if current_user.role == RoleType.Manager and user_in.role == RoleType.Admin:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Manager không có quyền tạo tài khoản Admin")

    assigned_facility = user_in.facility_id or current_user.facility_id

    new_user = User(
        username=user_in.username.strip(),
        password_hash=get_password_hash(user_in.password),
        full_name=user_in.full_name.strip() if user_in.full_name else None,
        role=user_in.role,
        facility_id=assigned_facility,
        phone_number=user_in.phone_number.strip() if user_in.phone_number else None,
        is_active=user_in.is_active,
        must_change_password=True
    )

    db.add(new_user)
    db.commit()
    db.refresh(new_user)

    fac = db.query(Facility).filter(Facility.id == new_user.facility_id).first() if new_user.facility_id else None

    return UserResponse(
        id=new_user.id,
        username=new_user.username,
        full_name=new_user.full_name,
        phone_number=new_user.phone_number,
        role=new_user.role,
        facility_id=new_user.facility_id,
        facility_name=fac.name if fac else None,
        is_active=new_user.is_active,
        must_change_password=new_user.must_change_password,
        created_at=new_user.created_at
    )


# =========================================================================
# 3. CẬP NHẬT THÔNG TIN TÀI KHOẢN (VAI TRÒ, CƠ SỞ, HỌ TÊN, SĐT)
# =========================================================================
@router.put("/{user_id}", response_model=UserResponse)
def update_user_info(
    user_id: int,
    payload: UserUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_management)
):
    user = db.query(User).filter(User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Không tìm thấy tài khoản người dùng")

    # Manager không được sửa tài khoản Admin hoặc gán quyền Admin cho ai đó
    if current_user.role == RoleType.Manager:
        if user.role == RoleType.Admin:
            raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Manager không có quyền chỉnh sửa tài khoản Admin")
        if payload.role == RoleType.Admin:
            raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Manager không thể gán quyền Admin")

    # Không thể tự hạ quyền hoặc tự khóa tài khoản của chính mình
    if user.id == current_user.id:
        if payload.is_active is not None and not payload.is_active:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Không thể tự khóa tài khoản của chính mình")
        if payload.role is not None and payload.role != current_user.role:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Không thể tự thay đổi vai trò của chính mình")

    if payload.full_name is not None:
        user.full_name = payload.full_name.strip()
    if payload.phone_number is not None:
        user.phone_number = payload.phone_number.strip()
    if payload.role is not None:
        user.role = payload.role
    if payload.facility_id is not None:
        user.facility_id = payload.facility_id
    if payload.is_active is not None:
        user.is_active = payload.is_active

    db.commit()
    db.refresh(user)

    fac = db.query(Facility).filter(Facility.id == user.facility_id).first() if user.facility_id else None

    return UserResponse(
        id=user.id,
        username=user.username,
        full_name=user.full_name,
        phone_number=user.phone_number,
        role=user.role,
        facility_id=user.facility_id,
        facility_name=fac.name if fac else None,
        is_active=user.is_active,
        must_change_password=user.must_change_password,
        created_at=user.created_at
    )


# =========================================================================
# 4. KHÓA / MỞ KHÓA TỪNG TÀI KHOẢN (TOGGLE LOCK)
# =========================================================================
@router.put("/{user_id}/toggle-lock", response_model=UserResponse)
def toggle_lock_user(
    user_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_management)
):
    user = db.query(User).filter(User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Không tìm thấy người dùng")

    if user.id == current_user.id:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Không thể tự khóa tài khoản của chính mình")

    # Bảo vệ tài khoản Admin Gốc
    if user.id == ROOT_ADMIN_ID:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Tài khoản Quản trị viên Gốc không thể bị khóa")

    # Manager không được khóa tài khoản Admin
    if current_user.role == RoleType.Manager and user.role == RoleType.Admin:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Bạn không có quyền khóa tài khoản Admin")

    user.is_active = not user.is_active
    db.commit()
    db.refresh(user)

    fac = db.query(Facility).filter(Facility.id == user.facility_id).first() if user.facility_id else None

    return UserResponse(
        id=user.id,
        username=user.username,
        full_name=user.full_name,
        phone_number=user.phone_number,
        role=user.role,
        facility_id=user.facility_id,
        facility_name=fac.name if fac else None,
        is_active=user.is_active,
        must_change_password=user.must_change_password,
        created_at=user.created_at
    )


# =========================================================================
# 5. KHÓA / MỞ KHÓA HÀNG LOẠT (BULK LOCK/UNLOCK)
# =========================================================================
@router.post("/bulk-lock", status_code=status.HTTP_200_OK)
def bulk_lock_users(
    payload: BulkLockRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_management)
):
    if not payload.user_ids:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Danh sách ID rỗng")

    # Loại trừ tài khoản chính mình và Admin Gốc khi khóa
    target_ids = [uid for uid in payload.user_ids if uid != current_user.id and uid != ROOT_ADMIN_ID]

    query = db.query(User).filter(User.id.in_(target_ids))

    # Manager không thể khóa Admin
    if current_user.role == RoleType.Manager:
        query = query.filter(User.role != RoleType.Admin)

    updated_count = query.update({User.is_active: payload.is_active}, synchronize_session=False)
    db.commit()

    return {
        "status": "Success",
        "action": "lock" if not payload.is_active else "unlock",
        "affected_count": updated_count
    }


# =========================================================================
# 6. XÓA TÀI KHOẢN ĐƠN LẺ & HÀNG LOẠT (CHỈ ROOT ADMIN ĐƯỢC XÓA ADMIN KHÁC)
# =========================================================================
@router.delete("/{user_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_user(
    user_id: int, 
    db: Session = Depends(get_db),
    current_user: User = Depends(require_management)
):
    user = db.query(User).filter(User.id == user_id).first()
    if not user: 
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Không tìm thấy người dùng")

    if user.id == current_user.id:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Không thể tự xóa tài khoản của chính mình")

    # 1. Tài khoản Admin Gốc (ID=1) tuyệt đối không ai được xóa
    if user.id == ROOT_ADMIN_ID:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN, 
            detail="Tài khoản Quản trị viên Gốc (Super Admin) được bảo vệ, không thể bị xóa"
        )

    # 2. Xóa tài khoản Admin: CHỈ DUY NHẤT Admin Gốc mới có quyền xóa Admin khác
    if user.role == RoleType.Admin:
        if not is_root_admin(current_user):
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN, 
                detail="Chỉ có Tài khoản Quản trị viên Gốc (Super Admin) mới có quyền xóa các tài khoản Admin khác"
            )

    # 3. Manager không thể xóa Admin
    if current_user.role == RoleType.Manager and user.role == RoleType.Admin:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Manager không được phép xóa tài khoản Admin")

    db.delete(user)
    db.commit()


@router.post("/bulk-delete", status_code=status.HTTP_200_OK)
def bulk_delete_users(
    payload: BulkActionRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_management)
):
    """
    Xóa hàng loạt tài khoản:
    - Loại trừ tài khoản đang đăng nhập và tài khoản Gốc (ID=1).
    - Nếu người gọi KHÔNG PHẢI là Admin Gốc: Tự động loại trừ toàn bộ tài khoản Admin.
    """
    if not payload.user_ids:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Danh sách ID rỗng")

    query = db.query(User).filter(
        User.id.in_(payload.user_ids),
        User.id != current_user.id,
        User.id != ROOT_ADMIN_ID
    )

    # Nếu không phải Admin Gốc -> Tuyệt đối không xóa bất cứ Admin nào
    if not is_root_admin(current_user):
        query = query.filter(User.role != RoleType.Admin)

    deleted_count = query.delete(synchronize_session=False)
    db.commit()

    return {
        "status": "Success",
        "deleted_count": deleted_count
    }


# =========================================================================
# 7. TIỆN ÍCH UX/UI: RESET MẬT KHẨU NHANH CHO NHÂN VIÊN
# =========================================================================
@router.put("/{user_id}/reset-password", status_code=status.HTTP_200_OK)
def reset_staff_password(
    user_id: int,
    payload: Optional[UserResetPassword] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_management)
):
    """
    Tiện ích Quản lý cấp lại mật khẩu nhanh khi nhân viên quên:
    - Mặc định: Reset về chính username/SĐT của tài khoản đó.
    - Đặt must_change_password = True để yêu cầu đổi mật khẩu ở lần đăng nhập tiếp theo.
    """
    user = db.query(User).filter(User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Không tìm thấy người dùng")

    if current_user.role == RoleType.Manager and user.role == RoleType.Admin:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Manager không có quyền reset mật khẩu Admin")

    default_new_pwd = payload.new_password if (payload and payload.new_password) else user.username
    user.password_hash = get_password_hash(default_new_pwd)
    user.must_change_password = True

    db.commit()

    return {
        "status": "Success",
        "message": f"Đã đặt lại mật khẩu cho tài khoản {user.username} thành công.",
        "default_password": default_new_pwd if (not payload or not payload.new_password) else "******"
    }


# =========================================================================
# HELPER: CHUẨN HÓA VAI TRÒ TỪ CHUỖI VIẾT TẮT TRONG EXCEL
# =========================================================================
def parse_role_from_string(role_str: Optional[str]) -> RoleType:
    if not role_str:
        return RoleType.Caregiver

    clean = str(role_str).strip().lower()

    if clean in ["dp", "đp"] or any(k in clean for k in ["điều phối", "dieu phoi", "coordinator"]):
        return RoleType.Coordinator
    if clean == "bs" or any(k in clean for k in ["bác sĩ", "bac si", "doctor"]):
        return RoleType.Doctor
    if clean == "ql" or any(k in clean for k in ["quản lý", "quan ly", "manager"]):
        return RoleType.Manager
    if clean == "admin" or any(k in clean for k in ["quản trị", "quan tri"]):
        return RoleType.Admin
    if clean == "nvcs" or any(k in clean for k in ["chăm sóc", "cham soc", "caregiver", "điều dưỡng", "y tá"]):
        return RoleType.Caregiver
    if clean == "bv" or any(k in clean for k in ["bảo vệ", "bao ve", "security"]):
        return RoleType.Security
    if any(k in clean for k in ["bếp", "bep", "kitchen", "cấp dưỡng"]):
        return RoleType.Kitchen
    if any(k in clean for k in ["tạp vụ", "tap vu", "janitor", "vệ sinh"]):
        return RoleType.Janitor
    if any(k in clean for k in ["người thân", "nguoi than", "relative", "gia đình"]):
        return RoleType.Relative

    return RoleType.Caregiver

# =========================================================================
# 8. IMPORT EXCEL & LỊCH SỬ TỔNG HỢP CỦA NHÂN VIÊN
# =========================================================================
@router.post(
    "/import-xlsx",
    status_code=status.HTTP_200_OK,
    summary="Import danh sách Nhân sự hàng loạt từ file Excel",
    description="""
    **Phân quyền & Kiểm tra Đa cơ sở:**
    - Admin / Manager Tổng (`facility_id is None`): Toàn quyền import cho tất cả cơ sở hoặc tạo nhân sự Toàn viện.
    - Manager Cơ sở X (`facility_id = X`):
      + Tự động BỎ QUA các dòng nhân sự thuộc cơ sở khác.
      + Tuyệt đối KHÔNG ĐƯỢC sửa thông tin nhân sự đã có sẵn của cơ sở khác hoặc Admin.
      + Dòng để trống cơ sở sẽ tự động gán vào Cơ sở X của Manager đó.
    """
)
async def import_staff_from_xlsx(
    file: UploadFile = File(..., description="File Excel danh sách nhân sự (.xlsx/.xls)"),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_management)
):
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST, 
            detail="Hệ thống chỉ chấp nhận file Excel định dạng .xlsx hoặc .xls"
        )

    try:
        contents = await file.read()
        wb = load_workbook(io.BytesIO(contents), data_only=True)
        ws = wb.active

        # 1. Bóc tách bản đồ các ô bị Merge ở Cột A (Cơ sở)
        merged_facility_map: Dict[int, str] = {}
        for rng in ws.merged_cells.ranges:
            if rng.min_col <= 1 <= rng.max_col:
                top_val = ws.cell(row=rng.min_row, column=rng.min_col).value
                if top_val is not None and str(top_val).strip() != "":
                    for row_idx in range(rng.min_row, rng.max_row + 1):
                        merged_facility_map[row_idx] = str(top_val).strip()

        facilities = db.query(Facility).all()
        inserted_count = 0
        updated_count = 0
        skipped_count = 0
        skipped_reasons: List[str] = []

        default_password_hash = get_password_hash("123456")

        # 2. Duyệt từng dòng dữ liệu từ hàng 2
        for row_idx in range(2, ws.max_row + 1):
            fullname_cell = ws.cell(row=row_idx, column=2).value
            phone_cell = ws.cell(row=row_idx, column=3).value
            role_cell = ws.cell(row=row_idx, column=4).value

            if not fullname_cell or not phone_cell:
                continue

            full_name = str(fullname_cell).strip()

            raw_phone = re.sub(r"[^\d]", "", str(phone_cell).strip())
            if len(raw_phone) == 9:
                raw_phone = "0" + raw_phone
            phone_str = raw_phone

            if not phone_str:
                continue

            # -------------------------------------------------------------
            # BƯỚC 1: XÁC ĐỊNH CƠ SỞ VÀ KIỂM TRA QUYỀN HẠN CỦA MANAGER
            # -------------------------------------------------------------
            if row_idx in merged_facility_map:
                fac_val = merged_facility_map[row_idx]
            else:
                raw_a = ws.cell(row=row_idx, column=1).value
                fac_val = str(raw_a).strip() if (raw_a is not None and str(raw_a).strip() != "") else None

            target_facility_id = None
            if fac_val:
                fac_str_lower = fac_val.lower()
                for f in facilities:
                    f_name_lower = f.name.lower()
                    if (f_name_lower == fac_str_lower or 
                        f"cs {fac_str_lower}" in f_name_lower or 
                        f"cơ sở {fac_str_lower}" in f_name_lower or
                        fac_str_lower in f_name_lower):
                        target_facility_id = f.id
                        break
            else:
                # Nếu file để trống:
                # - Admin/Manager Tổng -> facility_id = None (Toàn viện)
                # - Manager Cơ sở X -> tự động gán vào Cơ sở X
                target_facility_id = None if current_user.facility_id is None else current_user.facility_id

            # 🛡️ CHỐT CHẶN 1: Manager Cơ sở X cố tình import dòng thuộc Cơ sở Y khác
            if current_user.facility_id is not None:
                if target_facility_id is not None and target_facility_id != current_user.facility_id:
                    skipped_count += 1
                    skipped_reasons.append(f"Dòng {row_idx} ({full_name}): Thuộc cơ sở khác, bạn không có quyền import.")
                    continue
                # Luôn đảm bảo tài khoản tạo bởi Manager này thuộc đúng cơ sở của họ
                target_facility_id = current_user.facility_id

            # -------------------------------------------------------------
            # BƯỚC 2: XÁC ĐỊNH VAI TRÒ VÀ KHÓA BẢO VỆ ADMIN
            # -------------------------------------------------------------
            resolved_role = parse_role_from_string(str(role_cell) if role_cell else None)

            # 🛡️ CHỐT CHẶN 2: Manager không được cấp quyền Admin
            if current_user.role == RoleType.Manager and resolved_role == RoleType.Admin:
                resolved_role = RoleType.Caregiver

            # -------------------------------------------------------------
            # BƯỚC 3: KIỂM TRA TÀI KHOẢN TỒN TẠI VÀ CHỐNG SỬA ĐÈ CƠ SỞ KHÁC
            # -------------------------------------------------------------
            username = phone_str

            existing_user = db.query(User).filter(
                (User.username == username) | (User.phone_number == phone_str)
            ).first()

            if existing_user:
                # 🛡️ CHỐT CHẶN 3: Không cho Manager sửa tài khoản của Admin hoặc của Cơ sở khác
                if current_user.facility_id is not None:
                    if existing_user.facility_id != current_user.facility_id or existing_user.role == RoleType.Admin:
                        skipped_count += 1
                        skipped_reasons.append(f"Dòng {row_idx} ({full_name} - {phone_str}): Đã tồn tại nhưng thuộc quyền quản lý của Cơ sở khác hoặc là Admin.")
                        continue

                # Cập nhật thông tin hợp lệ
                existing_user.full_name = full_name
                existing_user.phone_number = phone_str
                existing_user.role = resolved_role
                if target_facility_id is not None or current_user.facility_id is None:
                    existing_user.facility_id = target_facility_id
                updated_count += 1
            else:
                # Tạo mới tài khoản hợp lệ
                new_user = User(
                    username=username,
                    password_hash=default_password_hash,
                    full_name=full_name,
                    role=resolved_role,
                    facility_id=target_facility_id,
                    phone_number=phone_str,
                    is_active=True,
                    must_change_password=True
                )
                db.add(new_user)
                inserted_count += 1

        db.commit()
        return {
            "status": "Success",
            "message": "Đồng bộ danh sách nhân sự từ Excel thành công!",
            "summary": {
                "total_created": inserted_count,
                "total_updated": updated_count,
                "total_skipped": skipped_count,
                "skipped_details": skipped_reasons[:10]  # Trả về tối đa 10 lý do nếu có dòng bị bỏ qua
            }
        }

    except Exception as e:
        db.rollback()
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, 
            detail=f"Lỗi cấu trúc hoặc đọc file Excel nhân sự: {str(e)}"
        )