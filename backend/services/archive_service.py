# services/archive_service.py
import io
import logging
import threading
from datetime import datetime, timedelta
from typing import Optional
import pandas as pd
import pytz
from sqlalchemy.orm import Session
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from database import SessionLocal
import models
from services.drive_service import upload_archive_file_to_drive

logger = logging.getLogger("archive_service")

_last_cleanup_date = None
_cleanup_lock = threading.Lock()


def format_excel_workbook(writer):
    """Trang trí file Excel theo giao diện chuẩn nghiệp vụ."""
    workbook = writer.book
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'), right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'), bottom=Side(style='thin', color='D9D9D9')
    )

    for sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align

        for col in worksheet.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = col[0].column_letter
            worksheet.column_dimensions[col_letter].width = max(max_len + 4, 12)
            for cell in col:
                if cell.row > 1:
                    cell.border = thin_border
                    cell.font = Font(name="Arial", size=9)


# =========================================================================
# 1. RÁC TỨC THỜI: DỌN SẠCH NONCE (QUÁ 10 PHÚT HOẶC ĐÃ DÙNG)
# =========================================================================
def purge_expired_nonces(db: Session) -> int:
    """Xóa bỏ vĩnh viễn nonces OTP/bảo mật quá 10 phút khỏi DB."""
    tz = pytz.timezone('Asia/Ho_Chi_Minh')
    threshold = datetime.now(tz) - timedelta(minutes=10)
    
    deleted = db.query(models.Nonce).filter(
        (models.Nonce.expires_at < threshold) | (models.Nonce.used == True)
    ).delete(synchronize_session=False)
    
    db.commit()
    if deleted > 0:
        logger.info(f"[CLEANUP NONCE]: Đã xóa {deleted} nonce rác.")
    return deleted


# =========================================================================
# 2. DỌN NHẬT KÝ ĐI TUẦN (INSPECTION LOGS > 30 NGÀY - PHÂN TÁCH THEO CƠ SỞ)
# =========================================================================
def archive_and_purge_facility_inspections(db: Session, facility: models.Facility, days: int = 30):
    tz = pytz.timezone('Asia/Ho_Chi_Minh')
    cutoff = datetime.now(tz) - timedelta(days=days)

    # Lấy logs thuộc cơ sở này thông qua Asset -> Room -> Zone -> Facility
    logs = db.query(
        models.InspectionLog,
        models.Shift.shift_date,
        models.Shift.shift_type,
        models.Asset.asset_name,
        models.Room.room_number,
        models.User.full_name
    ).join(models.Shift, models.InspectionLog.shift_id == models.Shift.id)\
     .join(models.Asset, models.InspectionLog.asset_id == models.Asset.id)\
     .join(models.Room, models.Asset.room_id == models.Room.id)\
     .join(models.Zone, models.Room.zone_id == models.Zone.id)\
     .outerjoin(models.User, models.InspectionLog.user_id == models.User.id)\
     .filter(
         models.Zone.facility_id == facility.id,
         models.InspectionLog.created_at < cutoff,
         models.Shift.status == "Submitted"
     ).all()

    if not logs:
        return

    data = [{
        "Mã Log": l.InspectionLog.id,
        "Cơ Sở": facility.name,
        "Ngày Ca": str(l.shift_date),
        "Ca Trực": l.shift_type,
        "Phòng": l.room_number or "N/A",
        "Tài Sản": l.asset_name,
        "Trạng Thái": l.InspectionLog.status,
        "Ghi Chú": l.InspectionLog.note,
        "Nhân Viên Kiểm": l.full_name or "N/A",
        "Link Ảnh (Drive)": l.InspectionLog.image_url,
        "Thời Gian Đi Tuần": l.InspectionLog.created_at.strftime("%d/%m/%Y %H:%M:%S") if l.InspectionLog.created_at else ""
    } for l in logs]

    df = pd.DataFrame(data)
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='INSPECTION_LOGS', index=False)
        format_excel_workbook(writer)

    output.seek(0)
    now_str = datetime.now(tz).strftime("%Y%m%d_%H%M%S")
    year_str = datetime.now(tz).strftime("%Y")
    filename = f"InspectionLogs_Truoc_{cutoff.strftime('%Y%m%d')}_{now_str}.xlsx"

    # Đẩy vào đúng folder CS_xxx/Archives/Inspection_Logs/2026/
    drive_link = upload_archive_file_to_drive(
        file_bytes=output.getvalue(),
        filename=filename,
        facility_name=facility.name,
        subfolder_path=["Inspection_Logs", year_str]
    )
    logger.info(f"[ARCHIVE SUCCESS] [{facility.name}] Inspection Logs -> {drive_link}")

    # Xóa DB sau khi upload thành công
    log_ids = [l.InspectionLog.id for l in logs]
    db.query(models.InspectionLog).filter(models.InspectionLog.id.in_(log_ids)).delete(synchronize_session=False)
    db.commit()
    logger.info(f"[PURGE DB SUCCESS] [{facility.name}] Đã giải phóng {len(log_ids)} inspection_logs.")


# =========================================================================
# 3. DỌN SINH HIỆU Y TẾ (VITAL SIGNS > 90 NGÀY - PHÂN TÁCH THEO CƠ SỞ)
# =========================================================================
def archive_and_purge_facility_vitals(db: Session, facility: models.Facility, days: int = 90):
    tz = pytz.timezone('Asia/Ho_Chi_Minh')
    cutoff = datetime.now(tz) - timedelta(days=days)

    records = db.query(
        models.VitalSignRecord,
        models.Elder.full_name.label("elder_name"),
        models.Room.room_number,
        models.User.full_name.label("staff_name")
    ).join(models.Elder, models.VitalSignRecord.elder_id == models.Elder.id)\
     .join(models.Room, models.Elder.room_id == models.Room.id)\
     .join(models.Zone, models.Room.zone_id == models.Zone.id)\
     .outerjoin(models.User, models.VitalSignRecord.measured_by == models.User.id)\
     .filter(
         models.Zone.facility_id == facility.id,
         models.VitalSignRecord.measured_at < cutoff
     ).all()

    if not records:
        return

    data = [{
        "Mã Bản Ghi": r.VitalSignRecord.id,
        "Cơ Sở": facility.name,
        "Phòng": r.room_number or "N/A",
        "Người Cao Tuổi": r.elder_name,
        "Huyết Áp": f"{r.VitalSignRecord.bp_systolic}/{r.VitalSignRecord.bp_diastolic}",
        "Mạch (bpm)": r.VitalSignRecord.pulse,
        "SpO2 (%)": r.VitalSignRecord.spo2,
        "Nhiệt Độ (°C)": r.VitalSignRecord.temperature,
        "Bất Thường": "Có" if r.VitalSignRecord.is_abnormal else "Không",
        "Ghi Chú": r.VitalSignRecord.notes,
        "Người Đo": r.staff_name or "N/A",
        "Thời Gian Đo": r.VitalSignRecord.measured_at.strftime("%d/%m/%Y %H:%M:%S") if r.VitalSignRecord.measured_at else ""
    } for r in records]

    df = pd.DataFrame(data)
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='VITAL_SIGNS', index=False)
        format_excel_workbook(writer)

    output.seek(0)
    now_str = datetime.now(tz).strftime("%Y%m%d_%H%M%S")
    year_str = datetime.now(tz).strftime("%Y")
    filename = f"VitalSigns_Truoc_{cutoff.strftime('%Y%m%d')}_{now_str}.xlsx"

    drive_link = upload_archive_file_to_drive(
        file_bytes=output.getvalue(),
        filename=filename,
        facility_name=facility.name,
        subfolder_path=["Medical_Vitals", year_str]
    )
    logger.info(f"[ARCHIVE SUCCESS] [{facility.name}] Vital Signs -> {drive_link}")

    rec_ids = [r.VitalSignRecord.id for r in records]
    db.query(models.VitalSignRecord).filter(models.VitalSignRecord.id.in_(rec_ids)).delete(synchronize_session=False)
    db.commit()
    logger.info(f"[PURGE DB SUCCESS] [{facility.name}] Đã giải phóng {len(rec_ids)} vital_sign_records.")


# =========================================================================
# 4. DỌN LOGIN LOGS & AUDIT LOGS (PHÂN THEO CƠ SỞ & GLOBAL)
# =========================================================================
def archive_and_purge_user_logs(db: Session, facility: Optional[models.Facility] = None, days: int = 30):
    """
    Lưu trữ Login Logs & Audit Logs:
    - Nếu có facility: Lọc logs của nhân viên thuộc facility_id đó.
    - Nếu facility=None: Lọc logs của Admin/Doctor toàn viện (facility_id is NULL) -> Đẩy vào System_Global.
    """
    tz = pytz.timezone('Asia/Ho_Chi_Minh')
    cutoff = datetime.now(tz) - timedelta(days=days)
    fac_name = facility.name if facility else "System_Global"
    fac_id = facility.id if facility else None

    # --- 1. XỬ LÝ LOGIN LOGS ---
    login_query = db.query(models.LoginLog, models.User.username, models.User.full_name)\
                    .join(models.User, models.LoginLog.user_id == models.User.id)\
                    .filter(models.LoginLog.login_time < cutoff)
    
    if fac_id is not None:
        login_query = login_query.filter(models.User.facility_id == fac_id)
    else:
        login_query = login_query.filter(models.User.facility_id.is_(None))

    login_logs = login_query.all()
    if login_logs:
        login_data = [{
            "ID": l.LoginLog.id,
            "Cơ Sở": fac_name,
            "Username": l.username,
            "Họ Tên": l.full_name,
            "IP Address": l.LoginLog.ip_address,
            "User Agent": l.LoginLog.user_agent,
            "Thời Gian": l.LoginLog.login_time.strftime("%d/%m/%Y %H:%M:%S") if l.LoginLog.login_time else ""
        } for l in login_logs]

        df_login = pd.DataFrame(login_data)
        out_login = io.BytesIO()
        with pd.ExcelWriter(out_login, engine='openpyxl') as writer:
            df_login.to_excel(writer, sheet_name='LOGIN_LOGS', index=False)
            format_excel_workbook(writer)

        out_login.seek(0)
        now_str = datetime.now(tz).strftime("%Y%m%d_%H%M%S")
        year_str = datetime.now(tz).strftime("%Y")
        f_name = f"LoginLogs_Truoc_{cutoff.strftime('%Y%m%d')}_{now_str}.xlsx"

        drive_link = upload_archive_file_to_drive(
            file_bytes=out_login.getvalue(),
            filename=f_name,
            facility_name=fac_name if facility else None,
            subfolder_path=["Login_Logs", year_str]
        )
        logger.info(f"[ARCHIVE SUCCESS] [{fac_name}] Login Logs -> {drive_link}")

        login_ids = [l.LoginLog.id for l in login_logs]
        db.query(models.LoginLog).filter(models.LoginLog.id.in_(login_ids)).delete(synchronize_session=False)
        db.commit()

    # --- 2. XỬ LÝ AUDIT LOGS (> 60 ngày) ---
    audit_cutoff = datetime.now(tz) - timedelta(days=60)
    audit_query = db.query(models.AuditLog, models.User.username, models.User.full_name)\
                    .outerjoin(models.User, models.AuditLog.actor_id == models.User.id)\
                    .filter(models.AuditLog.created_at < audit_cutoff)

    if fac_id is not None:
        audit_query = audit_query.filter(models.User.facility_id == fac_id)
    else:
        audit_query = audit_query.filter((models.User.facility_id.is_(None)) | (models.AuditLog.actor_id.is_(None)))

    audit_logs = audit_query.all()
    if audit_logs:
        audit_data = [{
            "Mã Log": a.AuditLog.id,
            "Cơ Sở": fac_name,
            "Người Thao Tác": f"{a.full_name} ({a.username})" if a.username else "Hệ Thống",
            "Hành Động": a.AuditLog.action,
            "Target ID": a.AuditLog.target_id,
            "Địa Chỉ IP": a.AuditLog.ip_address,
            "Chi Tiết Payload": a.AuditLog.payload,
            "Thời Gian": a.AuditLog.created_at.strftime("%d/%m/%Y %H:%M:%S") if a.AuditLog.created_at else ""
        } for a in audit_logs]

        df_audit = pd.DataFrame(audit_data)
        out_audit = io.BytesIO()
        with pd.ExcelWriter(out_audit, engine='openpyxl') as writer:
            df_audit.to_excel(writer, sheet_name='AUDIT_LOGS', index=False)
            format_excel_workbook(writer)

        out_audit.seek(0)
        now_str = datetime.now(tz).strftime("%Y%m%d_%H%M%S")
        year_str = datetime.now(tz).strftime("%Y")
        f_name = f"AuditLogs_Truoc_{audit_cutoff.strftime('%Y%m%d')}_{now_str}.xlsx"

        drive_link = upload_archive_file_to_drive(
            file_bytes=out_audit.getvalue(),
            filename=f_name,
            facility_name=fac_name if facility else None,
            subfolder_path=["Audit_Logs", year_str]
        )
        logger.info(f"[ARCHIVE SUCCESS] [{fac_name}] Audit Logs -> {drive_link}")

        audit_ids = [a.AuditLog.id for a in audit_logs]
        db.query(models.AuditLog).filter(models.AuditLog.id.in_(audit_ids)).delete(synchronize_session=False)
        db.commit()


# =========================================================================
# 5. HÀM ĐIỀU PHỐI CHÍNH: QUÉT TỪNG CƠ SỞ VÀ HỆ THỐNG CHUNG
# =========================================================================
def execute_full_system_cleanup():
    """Hàm tổng hợp quét dọn và lưu trữ dữ liệu theo từng cơ sở độc lập."""
    db = SessionLocal()
    try:
        logger.info("[MULTI-FACILITY CLEANUP]: Bắt đầu tiến trình lưu trữ dữ liệu theo Cơ sở...")
        
        # 1. Dọn Nonce tức thời
        purge_expired_nonces(db)

        # 2. Lấy danh sách tất cả Cơ sở hiện có
        facilities = db.query(models.Facility).all()

        # 3. Quét và lưu trữ độc lập cho từng Cơ sở
        for fac in facilities:
            logger.info(f"--- Đang xử lý lưu trữ cho Cơ sở: {fac.name} (ID: {fac.id}) ---")
            archive_and_purge_facility_inspections(db, fac, days=30)
            archive_and_purge_facility_vitals(db, fac, days=90)
            archive_and_purge_user_logs(db, facility=fac, days=30)

        # 4. Quét và lưu trữ logs cấp toàn viện (Admin/Doctor không gắn Cơ sở)
        logger.info("--- Đang xử lý lưu trữ cho System_Global ---")
        archive_and_purge_user_logs(db, facility=None, days=30)

        logger.info("[MULTI-FACILITY CLEANUP COMPLETE]: Hoàn tất chu trình dọn dẹp và lưu trữ đa cơ sở an toàn 100%!")
    except Exception as e:
        db.rollback()
        logger.error(f"[MULTI-FACILITY CLEANUP ERROR]: Lỗi dọn dẹp: {str(e)}")
    finally:
        db.close()


def trigger_archive_cleanup_in_background():
    """Kích hoạt tiến trình ngầm (Daemon Thread) - Throttle 1 lần/ngày."""
    global _last_cleanup_date
    tz = pytz.timezone('Asia/Ho_Chi_Minh')
    today_vn = datetime.now(tz).date()

    if _last_cleanup_date == today_vn:
        return

    if _cleanup_lock.acquire(blocking=False):
        try:
            if _last_cleanup_date != today_vn:
                _last_cleanup_date = today_vn
                thread = threading.Thread(target=execute_full_system_cleanup, daemon=True)
                thread.start()
                logger.info("[CLEANUP THREAD]: Đã kích hoạt Thread lưu trữ đa cơ sở ngầm hôm nay.")
        finally:
            _cleanup_lock.release()