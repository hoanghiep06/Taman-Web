# services/drive_service.py
import os
import io
import re
import pytz
import logging
from datetime import datetime, timedelta, date
from typing import List, Optional
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from core.config import settings
from google.auth.transport.requests import Request
from google.oauth2.credentials import Credentials
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseUpload

SCOPES = ['https://www.googleapis.com/auth/drive']
ROOT_FOLDER_ID = settings.GOOGLE_DRIVE_ROOT_FOLDER_ID

creds = Credentials(
    token=None,
    refresh_token=settings.GOOGLE_REFRESH_TOKEN,
    token_uri="https://oauth2.googleapis.com/token",
    client_id=settings.GOOGLE_CLIENT_ID,
    client_secret=settings.GOOGLE_CLIENT_SECRET,
    scopes=SCOPES
)

def get_drive_service():
    """Lazy Loading & Auto-Recovery kết nối Google Drive API."""
    try:
        if not creds or not creds.valid:
            creds.refresh(Request())
        return build('drive', 'v3', credentials=creds)
    except Exception as e:
        logging.error(f"[DRIVE CONNECT FATAL]: Lỗi API Drive: {str(e)}")
        raise Exception("Google Drive API chưa sẵn sàng.")

def sanitize_filename(name: str) -> str:
    """Loại bỏ ký tự đặc biệt, thay khoảng trắng bằng gạch dưới."""
    name = name.replace(" ", "_").replace("'", "")
    name = re.sub(r'[\\/*?:"<>|]', "", name)
    return name

def get_or_create_folder(folder_name: str, parent_id: str) -> str:
    """Tìm thư mục theo tên trong thư mục cha. Nếu chưa có thì tự động tạo mới."""
    service = get_drive_service()
    safe_name = folder_name.replace("'", "\\'")
    query = f"mimeType='application/vnd.google-apps.folder' and '{parent_id}' in parents and name='{safe_name}' and trashed=false"
    results = service.files().list(q=query, spaces='drive', fields='files(id, name)').execute()
    items = results.get('files', [])
    
    if items:
        return items[0]['id']
        
    folder_metadata = {
        'name': folder_name,
        'mimeType': 'application/vnd.google-apps.folder',
        'parents': [parent_id]
    }
    folder = service.files().create(body=folder_metadata, fields='id').execute()
    return folder.get('id')


# =========================================================================
# 1. ẢNH KIỂM KÊ TÀI SẢN (Chức Năng -> Cơ Sở -> Phòng)
# =========================================================================
def upload_image_to_drive(file_bytes: bytes, facility_name: str, shift_date: str, shift_type: str, room_number: str, asset_names: List[str]) -> str:
    service = get_drive_service()

    # 1. Thư mục Chức năng và Cơ sở
    func_folder_id = get_or_create_folder("Anh_Kiem_Ke", ROOT_FOLDER_ID)
    fac_folder_id = get_or_create_folder(sanitize_filename(facility_name), func_folder_id)
    
    # 2. Thư mục Ngày (Định dạng DD-MM-YYYY)
    try:
        dt_obj = datetime.strptime(str(shift_date).strip(), "%Y-%m-%d")
        date_folder_name = dt_obj.strftime("%d-%m-%Y")
    except Exception:
        date_folder_name = str(shift_date).replace("-", "")

    date_folder_id = get_or_create_folder(date_folder_name, fac_folder_id)

    # 3. Thư mục Ca (Ca_Sang / Ca_Toi)
    shift_folder_name = f"Ca_{shift_type}"
    shift_folder_id = get_or_create_folder(shift_folder_name, date_folder_id)

    # 4. Thư mục Phòng
    room_folder_id = get_or_create_folder(f"Phong_{sanitize_filename(room_number)}", shift_folder_id)

    # Đặt tên file ảnh
    sanitized_names = [sanitize_filename(name) for name in asset_names]
    assets_prefix = "__".join(sanitized_names)[:100]
    filename = f"{assets_prefix}_{int(datetime.now().timestamp())}.jpg"
    
    # Upload
    file_metadata = {'name': filename, 'parents': [room_folder_id]}
    media = MediaIoBaseUpload(io.BytesIO(file_bytes), mimetype='image/jpeg', resumable=True)
    file = service.files().create(body=file_metadata, media_body=media, fields='id, webViewLink').execute()
    
    return file.get('webViewLink')


# =========================================================================
# 2. ẢNH TOA THUỐC (Được dùng trong hàm Upload toa thuốc)
# =========================================================================
def upload_prescription_to_drive(file_bytes: bytes, facility_name: str, elder_name: str) -> str:
    service = get_drive_service()

    facility_folder_id = get_or_create_folder(sanitize_filename(facility_name), ROOT_FOLDER_ID)
    health_folder_id = get_or_create_folder("Health", facility_folder_id)
    prescriptions_folder_id = get_or_create_folder("Prescriptions", health_folder_id)

    date_str = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"ToaThuoc_{sanitize_filename(elder_name)}_{date_str}.jpg"

    file_metadata = {'name': filename, 'parents': [prescriptions_folder_id]}
    media = MediaIoBaseUpload(io.BytesIO(file_bytes), mimetype='image/jpeg', resumable=True)
    file = service.files().create(body=file_metadata, media_body=media, fields='id, webViewLink').execute()

    return file.get('webViewLink')


# =========================================================================
# 3. BÁO CÁO CA TRỰC (Chức Năng -> Cơ Sở -> Ngày)
# =========================================================================
def upload_shift_handover_report_to_drive(file_bytes: bytes, facility_name: str, shift_date_str: str, filename: str) -> str:
    service = get_drive_service()
    
    # Phân cấp: Bao_Cao_Ca_Truc -> Tên Cơ Sở -> DD-MM-YYYY
    func_folder_id = get_or_create_folder("Bao_Cao_Ca_Truc", ROOT_FOLDER_ID)
    fac_folder_id = get_or_create_folder(sanitize_filename(facility_name), func_folder_id)
    
    try:
        dt_obj = datetime.strptime(str(shift_date_str).strip(), "%Y-%m-%d")
        date_folder_name = dt_obj.strftime("%d-%m-%Y")
    except Exception:
        date_folder_name = str(shift_date_str)

    date_folder_id = get_or_create_folder(date_folder_name, fac_folder_id)
    media = MediaIoBaseUpload(io.BytesIO(file_bytes), mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', resumable=True)

    safe_filename = filename.replace("'", "\\'")
    query = f"name='{safe_filename}' and '{date_folder_id}' in parents and trashed=false"
    results = service.files().list(q=query, spaces='drive', fields='files(id, name)').execute()
    existing_files = results.get('files', [])

    if existing_files:
        file = service.files().update(fileId=existing_files[0]['id'], media_body=media, fields='id, webViewLink').execute()
    else:
        file_metadata = {'name': filename, 'parents': [date_folder_id]}
        file = service.files().create(body=file_metadata, media_body=media, fields='id, webViewLink').execute()

    return file.get('webViewLink')


# =========================================================================
# 4. BACKUP HỆ THỐNG & ARCHIVE CHUYÊN SÂU
# =========================================================================
def upload_db_backup_to_drive(file_bytes: bytes, filename: str) -> str:
    """🌟 Đã đổi lại đúng tên cũ để tương thích với scheduler.py"""
    service = get_drive_service()
    func_folder_id = get_or_create_folder("Backup_He_Thong", ROOT_FOLDER_ID)
    file_metadata = {'name': filename, 'parents': [func_folder_id]}
    media = MediaIoBaseUpload(io.BytesIO(file_bytes), mimetype='text/plain', resumable=True)
    file = service.files().create(body=file_metadata, media_body=media, fields='id, webViewLink').execute()
    return file.get('webViewLink')


def upload_archive_file_to_drive(
    file_bytes: bytes, 
    filename: str, 
    facility_name: Optional[str],
    subfolder_path: List[str]
) -> str:
    """Tải file lưu trữ lên Drive theo cây thư mục đa cơ sở: Root -> Tên_Cơ_Sở -> Archives -> Category -> Year -> filename"""
    service = get_drive_service()
    root_subfolder_name = sanitize_filename(facility_name) if facility_name else "System_Global"
    facility_folder_id = get_or_create_folder(root_subfolder_name, ROOT_FOLDER_ID)
    archives_folder_id = get_or_create_folder("Archives", facility_folder_id)
    
    current_parent_id = archives_folder_id
    for folder_name in subfolder_path:
        current_parent_id = get_or_create_folder(folder_name, current_parent_id)
        
    file_metadata = {'name': filename, 'parents': [current_parent_id]}
    media = MediaIoBaseUpload(
        io.BytesIO(file_bytes),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        resumable=True
    )
    file = service.files().create(body=file_metadata, media_body=media, fields='id, webViewLink').execute()
    return file.get('webViewLink')


def cleanup_old_db_backups(keep_count: int = 10):
    """Xóa các bản backup cũ vượt quá ngưỡng keep_count."""
    try:
        service = get_drive_service()
        backup_base_id = get_or_create_folder("Backup_He_Thong", ROOT_FOLDER_ID)
        query = f"'{backup_base_id}' in parents and trashed = false"
        results = service.files().list(
            q=query,
            fields="files(id, name, modifiedTime)",
            orderBy="modifiedTime desc",
            pageSize=100
        ).execute()
        
        files = results.get('files', [])
        if len(files) > keep_count:
            files_to_delete = files[keep_count:]
            for old_file in files_to_delete:
                service.files().delete(fileId=old_file['id']).execute()
                logging.info(f"[DRIVE RETENTION]: Đã xóa bản backup cũ: {old_file['name']}")
    except Exception as e:
        logging.error(f"[DRIVE RETENTION ERROR]: {str(e)}")


def list_db_backups_from_drive():
    """Truy vấn danh sách toàn bộ các file backup (.sql) trên Drive."""
    service = get_drive_service()
    backup_base_id = get_or_create_folder("Backup_He_Thong", ROOT_FOLDER_ID)
    query = f"'{backup_base_id}' in parents and trashed = false"
    
    results = service.files().list(
        q=query,
        fields="files(id, name, size, modifiedTime)",
        orderBy="modifiedTime desc",
        pageSize=50
    ).execute()
    
    return results.get('files', [])


# =========================================================================
# 5. DỌN DẸP RÁC ẢNH THEO NGÀY (CRONJOB)
# =========================================================================
def cleanup_old_drive_folders(days: int = 7):
    """Quét dẹp ảnh kiểm kê cũ dựa vào tên thư mục YYYYMMDD"""
    try:
        service = get_drive_service()
        tz = pytz.timezone('Asia/Ho_Chi_Minh')
        cutoff_date = (datetime.now(tz) - timedelta(days=days)).date()

        facility_query = f"'{ROOT_FOLDER_ID}' in parents and mimeType='application/vnd.google-apps.folder' and trashed=false"
        facility_results = service.files().list(q=facility_query, fields="files(id, name)").execute()
        facilities = facility_results.get('files', [])

        deleted_count = 0

        for fac in facilities:
            if fac['name'] in ["Backup_He_Thong", "tmp", "Anh_Kiem_Ke", "Bao_Cao_Ca_Truc", "Suc_Khoe_Hang_Ngay"]:
                continue

            insp_query = f"name='InspectionImage' and '{fac['id']}' in parents and mimeType='application/vnd.google-apps.folder' and trashed=false"
            insp_results = service.files().list(q=insp_query, fields="files(id, name)").execute()
            insp_folders = insp_results.get('files', [])

            for insp_folder in insp_folders:
                insp_folder_id = insp_folder['id']
                date_folders_query = f"'{insp_folder_id}' in parents and mimeType='application/vnd.google-apps.folder' and trashed=false"
                date_results = service.files().list(q=date_folders_query, fields="files(id, name)").execute()
                date_folders = date_results.get('files', [])

                for df in date_folders:
                    folder_name = df['name']
                    if len(folder_name) == 8 and folder_name.isdigit():
                        try:
                            folder_date = datetime.strptime(folder_name, "%Y%m%d").date()
                            if folder_date < cutoff_date:
                                service.files().delete(fileId=df['id']).execute()
                                deleted_count += 1
                        except ValueError:
                            continue

        logging.info(f"[DRIVE CLEANUP COMPLETE]: Hoàn tất dọn dẹp. Tổng số thư mục ngày đã xóa: {deleted_count}")
    except Exception as e:
        logging.error(f"[DRIVE CLEANUP ERROR]: Lỗi khi dọn dẹp ảnh kiểm kê: {e}")


# =========================================================================
# 6. HELPERS XỬ LÝ ID URL & TẢI FILE TỪ CLOUD
# =========================================================================
def extract_file_id_from_url(url: str) -> Optional[str]:
    if not url: return None
    url_str = str(url).strip()
    if re.match(r'^[a-zA-Z0-9_-]{20,}$', url_str): return url_str
    match = re.search(r'/d/([a-zA-Z0-9_-]+)', url_str)
    if match: return match.group(1)
    match = re.search(r'id=([a-zA-Z0-9_-]+)', url_str)
    if match: return match.group(1)
    return None

def download_file_bytes_from_drive(file_id: str) -> bytes:
    service = get_drive_service()
    return service.files().get_media(fileId=file_id).execute()


# =========================================================================
# 7. SỨC KHỎE HẰNG NGÀY (CÓ CƠ CHẾ TÌM DÒNG CŨ ĐỂ GHI ĐÈ THÔNG MINH)
# =========================================================================
def append_elder_health_log_to_drive(facility_name: str, elder_name: str, log_data: dict):
    service = get_drive_service()
    
    func_folder_id = get_or_create_folder("Suc_Khoe_Hang_Ngay", ROOT_FOLDER_ID)
    fac_folder_id = get_or_create_folder(sanitize_filename(facility_name), func_folder_id)
    
    filename = f"{sanitize_filename(elder_name)}.xlsx"
    query = f"name='{filename}' and '{fac_folder_id}' in parents and trashed=false"
    results = service.files().list(q=query, spaces='drive', fields='files(id, name)').execute()
    existing_files = results.get('files', [])

    if existing_files:
        file_id = existing_files[0]['id']
        file_bytes = service.files().get_media(fileId=file_id).execute()
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes))
        ws = wb.active
    else:
        file_id = None
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Nhat Ky Suc Khoe"
        
        # Cấu trúc Cột Mới Y Hệt Hình Ảnh
        headers = ["Ngày Giờ", "Loại Ghi Nhận", "Huyết Áp", "Mạch", "Nhiệt độ", "SpO2", "Cân nặng", "Người Ghi Nhận", "Ghi Chú đo chỉ số", "Ghi Chú từ báo cáo giao ca"]
        ws.append(headers)
        
        header_fill = PatternFill(start_color="285A82", end_color="285A82", fill_type="solid")
        header_font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
        center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            
        ws.column_dimensions["A"].width = 16
        ws.column_dimensions["B"].width = 22
        ws.column_dimensions["C"].width = 11
        ws.column_dimensions["D"].width = 8
        ws.column_dimensions["E"].width = 10
        ws.column_dimensions["F"].width = 8
        ws.column_dimensions["G"].width = 12
        ws.column_dimensions["H"].width = 20
        ws.column_dimensions["I"].width = 35
        ws.column_dimensions["J"].width = 40

    is_update = log_data.get("is_update", False)
    date_key = log_data.get("date_key", "")
    base_type = log_data.get("base_type", "")
    target_row = ws.max_row + 1

    # NẾU LÀ CẬP NHẬT: Quét ngược từ dưới lên để tìm dòng tương ứng (cùng ngày & cùng loại) để ghi đè
    if is_update and date_key and base_type:
        for r_idx in range(ws.max_row, 1, -1):
            c_time = str(ws.cell(row=r_idx, column=1).value or "")
            c_type = str(ws.cell(row=r_idx, column=2).value or "")
            if date_key in c_time and base_type in c_type:
                target_row = r_idx
                break

    # Đổ dữ liệu vào đúng Target Row
    ws.cell(row=target_row, column=1, value=log_data.get("time", ""))
    ws.cell(row=target_row, column=2, value=log_data.get("type", ""))
    ws.cell(row=target_row, column=3, value=log_data.get("bp", ""))
    ws.cell(row=target_row, column=4, value=log_data.get("pulse", ""))
    ws.cell(row=target_row, column=5, value=log_data.get("temp", ""))
    ws.cell(row=target_row, column=6, value=log_data.get("spo2", ""))
    ws.cell(row=target_row, column=7, value=log_data.get("weight", ""))
    ws.cell(row=target_row, column=8, value=log_data.get("staff", ""))
    ws.cell(row=target_row, column=9, value=log_data.get("vital_note", ""))
    ws.cell(row=target_row, column=10, value=log_data.get("shift_note", ""))
    
    # Cấu hình Formatting Từng Ô (Chỉ đỏ những ô vượt ngưỡng)
    danger_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
    danger_font = Font(name="Arial", size=10, color="C00000", bold=True)
    normal_font = Font(name="Arial", size=10)
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'), right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'), bottom=Side(style='thin', color='D9D9D9')
    )

    flags = log_data.get("abnormal_flags", {})

    for col_idx in range(1, 11):
        cell = ws.cell(row=target_row, column=col_idx)
        cell.border = thin_border
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        cell.font = normal_font
        cell.fill = PatternFill(fill_type=None) # Xóa sạch màu cũ nếu là cập nhật đè

        is_danger = False
        if col_idx == 3 and flags.get("bp"): is_danger = True
        if col_idx == 4 and flags.get("pulse"): is_danger = True
        if col_idx == 5 and flags.get("temp"): is_danger = True
        if col_idx == 6 and flags.get("spo2"): is_danger = True
        if col_idx == 10 and flags.get("shift_note"): is_danger = True
        if col_idx == 9 and flags.get("any_vital"): is_danger = True 

        if is_danger:
            cell.fill = danger_fill
            cell.font = danger_font

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    media = MediaIoBaseUpload(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', resumable=True)

    if file_id:
        service.files().update(fileId=file_id, media_body=media).execute()
        logging.info(f"[HEALTH LOG APPEND]: Đã ghi chú/cập nhật đè vào {filename}")
    else:
        file_metadata = {'name': filename, 'parents': [fac_folder_id]}
        service.files().create(body=file_metadata, media_body=media).execute()
        logging.info(f"[HEALTH LOG CREATE]: Đã khởi tạo file {filename}")