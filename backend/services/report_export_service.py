import io
import re
import logging
import pytz
from datetime import datetime, date, timedelta
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session, joinedload

from database import SessionLocal
import models
from core.constants import VITAL_LIMITS
from services.drive_service import upload_shift_handover_report_to_drive, sanitize_filename

logger = logging.getLogger("report_export_service")


def export_and_upload_shift_report_background(report_id: int):
    db: Session = SessionLocal()
    try:
        # Load đầy đủ thông tin báo cáo kèm Cơ sở và Điều phối viên
        report = db.query(models.ShiftReport).options(
            joinedload(models.ShiftReport.facility),
            joinedload(models.ShiftReport.coordinator)
        ).filter(models.ShiftReport.id == report_id).first()

        if not report:
            logger.warning(f"Background task aborted. Report ID {report_id} not found.")
            return

        facility_name = report.facility.name if report.facility else "Cơ Sở Chưa Xác Định"
        coordinator_name = report.coordinator.full_name if report.coordinator else "N/A"
        shift_type_label = "Ca Sáng (08:00 - 19:00)" if report.shift_type == "Sang" else "Ca Tối (20:00 - 07:00)"
        shift_date_str = str(report.shift_date)

        # =========================================================================================
        # 1. TRUY VẤN VÀ ĐỒNG BỘ SINH HIỆU BẰNG KHUNG THỜI GIAN MỞ RỘNG (CHỐNG MẤT DỮ LIỆU CA TỐI)
        # =========================================================================================
        vn_tz = pytz.timezone('Asia/Ho_Chi_Minh')
        
        # Mở rộng biên độ thời gian an toàn (Safe Time Bounds) để cover hết các trường hợp nhập trễ
        # Dùng ngày báo cáo (report.shift_date) làm mốc gốc
        if report.shift_type == "Sang":
            # Quét sinh hiệu từ 06:00 sáng đến 19:30 tối
            local_start = vn_tz.localize(datetime.combine(report.shift_date, datetime.min.time().replace(hour=6, minute=0))) 
            local_end = vn_tz.localize(datetime.combine(report.shift_date, datetime.min.time().replace(hour=19, minute=30))) 
        else:
            # Ca Tối vắt ngang 2 ngày -> Lấy từ 19:00 hôm nay đến 08:00 sáng hôm sau
            local_start = vn_tz.localize(datetime.combine(report.shift_date, datetime.min.time().replace(hour=19, minute=0))) 
            # Dùng timedelta để tiến thẳng tới ngày hôm sau an toàn
            local_end = local_start + timedelta(hours=13) 
            
        # Convert về UTC để so khớp với DB PostgreSQL (nơi lưu `measured_at` theo chuẩn UTC)
        utc_start = local_start.astimezone(pytz.utc).replace(tzinfo=None)
        utc_end = local_end.astimezone(pytz.utc).replace(tzinfo=None)

        vitals = db.query(models.VitalSignRecord, models.Elder.full_name, models.Room.room_number, models.User.full_name.label("staff_name"))\
                   .join(models.Elder, models.VitalSignRecord.elder_id == models.Elder.id)\
                   .join(models.Room, models.Elder.room_id == models.Room.id)\
                   .join(models.Zone, models.Room.zone_id == models.Zone.id)\
                   .outerjoin(models.User, models.VitalSignRecord.measured_by == models.User.id)\
                   .filter(
                       models.Zone.facility_id == report.facility_id,
                       models.VitalSignRecord.shift_type == report.shift_type,
                       models.VitalSignRecord.measured_at >= utc_start,
                       models.VitalSignRecord.measured_at <= utc_end
                   ).order_by(models.Room.room_number, models.Elder.full_name).all()

        # =========================================================================================
        # 2. TẠO TÓM TẮT CẢNH BÁO TỰ ĐỘNG THÔNG MINH CHO SHEET 1
        # =========================================================================================
        alerts_text = []
        for v, elder_name, room_num, staff_name in vitals:
            # Khôi phục cờ gốc: Có sự cố sức khỏe (từ DB) hoặc có Ghi chú
            has_abnormal = bool(v.is_abnormal)
            staff_note = v.notes.strip() if v.notes else ""
            has_note = bool(staff_note)
            
            if has_abnormal or has_note:
                issue_details = []
                
                # Check các chỉ số chạm/vượt giới hạn
                if v.spo2 and v.spo2 < VITAL_LIMITS["SPO2_WARNING"]: 
                    issue_details.append(f"SpO2 thấp ({v.spo2}%)")
                if v.temperature and v.temperature >= VITAL_LIMITS["TEMP_FEVER"]: 
                    issue_details.append(f"Sốt ({v.temperature}°C)")
                
                # Check Huyết áp (Dùng logic OR giống FE: Nếu bất cứ chiều nào lệch là báo lỗi ngay)
                bp_sys = v.bp_systolic
                bp_dia = v.bp_diastolic
                bp_abnormal = False
                if bp_sys and (bp_sys > VITAL_LIMITS["BP_SYSTOLIC_HIGH"] or bp_sys < VITAL_LIMITS["BP_SYSTOLIC_LOW"]):
                    bp_abnormal = True
                if bp_dia and (bp_dia > VITAL_LIMITS["BP_DIASTOLIC_HIGH"] or bp_dia < VITAL_LIMITS["BP_DIASTOLIC_LOW"]):
                    bp_abnormal = True
                
                if bp_abnormal:
                    # Gộp chung vào 1 cảnh báo để dễ đọc
                    if bp_sys and bp_sys > VITAL_LIMITS["BP_SYSTOLIC_HIGH"]:
                        issue_details.append(f"Huyết áp cao ({bp_sys}/{bp_dia})")
                    elif bp_sys and bp_sys < VITAL_LIMITS["BP_SYSTOLIC_LOW"]:
                        issue_details.append(f"Huyết áp thấp ({bp_sys}/{bp_dia})")
                    else:
                        issue_details.append(f"Huyết áp bất thường ({bp_sys}/{bp_dia})")
                
                # Check Mạch đập
                if v.pulse and (v.pulse > VITAL_LIMITS["PULSE_FAST"] or v.pulse < VITAL_LIMITS["PULSE_SLOW"]): 
                    issue_details.append(f"Mạch ({v.pulse} bpm)")
                
                # Định dạng string đầu ra
                detail_str = " - ".join(issue_details)
                
                if detail_str and staff_note:
                    alerts_text.append(f"Phòng {room_num} • {elder_name}: {detail_str}\n  >> (Ghi chú NV: {staff_note})")
                elif detail_str:
                    alerts_text.append(f"Phòng {room_num} • {elder_name}: {detail_str}")
                elif staff_note:
                    alerts_text.append(f"Phòng {room_num} • {elder_name}: (Ghi chú NV: {staff_note})")

        # Nối tất cả các dòng cảnh báo lại
        alerts_summary = "\n\n".join(alerts_text) if alerts_text else "Ca trực bình thường, không ghi nhận diễn biến đặc biệt (Tất cả các Cụ đều ổn định)."

        # =========================================================================================
        # 3. KHỞI TẠO BỘ CÔNG CỤ EXCEL & ĐỊNH DẠNG (STYLE)
        # =========================================================================================
        wb = openpyxl.Workbook()
        
        # Bảng màu
        title_font = Font(name="Arial", size=13, bold=True, color="1F4E78")
        section_font = Font(name="Arial", size=10, bold=True, color="1F4E78")
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        table_header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        table_header_font = Font(name="Arial", size=9, bold=True, color="1F4E78")
        header_font = Font(name="Arial", size=9, bold=True, color="FFFFFF")
        bold_font = Font(name="Arial", size=9, bold=True)
        normal_font = Font(name="Arial", size=9)
        
        # Formats cảnh báo
        danger_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
        danger_font = Font(name="Arial", size=9, color="C00000", bold=True)
        note_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        note_font = Font(name="Arial", size=9, color="B45F06", bold=True)
        
        # Alignment & Borders
        center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        left_align_wrap = Alignment(horizontal="left", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style='thin', color='D9D9D9'), right=Side(style='thin', color='D9D9D9'),
            top=Side(style='thin', color='D9D9D9'), bottom=Side(style='thin', color='D9D9D9')
        )

        # =========================================================================================
        # SHEET 1: BIÊN BẢN BÀN GIAO CA TRỰC
        # =========================================================================================
        ws1 = wb.active
        ws1.title = "BÀN GIAO CA TRỰC"
        ws1.views.sheetView[0].showGridLines = True

        ws1.merge_cells("A1:F1")
        ws1["A1"] = "BIÊN BẢN BÁO CÁO GIAO CA TRỰC - VIỆN DƯỠNG LÃO TÂM AN"
        ws1["A1"].font = title_font
        ws1["A1"].alignment = center_align
        ws1.row_dimensions[1].height = 28

        # Chuyển đổi timestamp tạo báo cáo về giờ Việt Nam (Hiển thị Header)
        created_at_vn = ""
        if report.created_at:
            dt_utc = pytz.utc.localize(report.created_at) if report.created_at.tzinfo is None else report.created_at
            created_at_vn = dt_utc.astimezone(vn_tz).strftime("%d/%m/%Y %H:%M:%S")

        meta_info = [
            ("Cơ Sở:", facility_name, "Ca Trực Bàn Giao:", shift_type_label),
            ("Ngày Ca Trực:", report.shift_date.strftime("%d/%m/%Y"), "Người Lập Báo Cáo:", coordinator_name),
            ("Thời Gian Nộp:", created_at_vn, "Trạng Thái:", "Đã xác nhận giao ca (Submitted)")
        ]

        for r_idx, (l1, v1, l2, v2) in enumerate(meta_info, start=3):
            ws1.cell(row=r_idx, column=1, value=l1).font = bold_font
            ws1.cell(row=r_idx, column=2, value=v1).font = normal_font
            ws1.cell(row=r_idx, column=4, value=l2).font = bold_font
            ws1.cell(row=r_idx, column=5, value=v2).font = normal_font
            ws1.row_dimensions[r_idx].height = 20

        # ---- MỤC 1: TÓM TẮT VẤN ĐỀ VÀ CHỈ SỐ BẤT THƯỜNG ----
        ws1.cell(row=7, column=1, value="1. CÁC CHỈ SỐ SỨC KHỎE CẦN LƯU Ý:").font = section_font
        ws1.merge_cells("A8:F8")
        ws1["A8"] = alerts_summary
        
        # Nếu mảng cảnh báo có phần tử -> Đánh dấu chữ đỏ để nổi bật
        if len(alerts_text) > 0:
            ws1["A8"].font = danger_font
            ws1["A8"].fill = danger_fill
        else:
            ws1["A8"].font = normal_font
        ws1["A8"].alignment = left_align_wrap
        
        # Auto-height (mỗi dòng gán 16 height)
        num_lines = alerts_summary.count('\n') + 1
        ws1.row_dimensions[8].height = max(25, num_lines * 16) 

        # ---- MỤC 2: HƯỚNG XỬ LÝ ----
        ws1.cell(row=10, column=1, value="2. LƯU Ý & HƯỚNG XỬ LÝ CA TIẾP THEO:").font = section_font
        ws1.merge_cells("A11:F12")
        ws1["A11"] = report.handover_notes or "Bàn giao ca bình thường."
        ws1["A11"].font = note_font if report.handover_notes else normal_font
        if report.handover_notes:
            ws1["A11"].fill = note_fill
        ws1["A11"].alignment = left_align_wrap

        # ---- MỤC 3: BẢNG DIỄN BIẾN CHI TIẾT TỪNG CỤ ----
        ws1.cell(row=14, column=1, value="3. DIỄN BIẾN CHI TIẾT TỪNG CỤ:").font = section_font
        
        ws1.cell(row=15, column=1, value="STT").fill = table_header_fill
        ws1.cell(row=15, column=1).font = table_header_font
        ws1.cell(row=15, column=1).alignment = center_align
        ws1.cell(row=15, column=1).border = thin_border

        ws1.cell(row=15, column=2, value="Họ Và Tên Cụ").fill = table_header_fill
        ws1.cell(row=15, column=2).font = table_header_font
        ws1.cell(row=15, column=2).alignment = Alignment(horizontal="left", vertical="center")
        ws1.cell(row=15, column=2).border = thin_border

        ws1.merge_cells("C15:F15")
        ws1.cell(row=15, column=3, value="Diễn Biến & Ghi Chú Chi Tiết Trong Ca").fill = table_header_fill
        ws1.cell(row=15, column=3).font = table_header_font
        ws1.cell(row=15, column=3).alignment = Alignment(horizontal="left", vertical="center")
        for col_c in range(3, 7):
            ws1.cell(row=15, column=col_c).border = thin_border

        raw_lines = [l.strip() for l in (report.elder_descriptions or "").split("\n") if l.strip()]
        current_r = 16

        if not raw_lines:
            ws1.cell(row=16, column=1, value="-").alignment = center_align
            ws1.cell(row=16, column=2, value="Tất cả các Cụ").alignment = Alignment(horizontal="left", vertical="center")
            ws1.merge_cells("C16:F16")
            ws1.cell(row=16, column=3, value="Sinh hoạt và sức khỏe ổn định trong suốt ca trực.")
            for c_i in range(1, 7): ws1.cell(row=16, column=c_i).border = thin_border
            current_r = 17
        else:
            for idx, line in enumerate(raw_lines, start=1):
                ws1.row_dimensions[current_r].height = 22
                
                # Split tên Cụ
                if ":" in line:
                    parts = line.split(":", 1)
                    e_name = re.sub(r"^\d+\.\s*", "", parts[0]).strip()
                    e_note = parts[1].strip()
                else:
                    e_name = f"Mục {idx}"
                    e_note = line

                c1 = ws1.cell(row=current_r, column=1, value=idx)
                c1.alignment = center_align
                c1.border = thin_border
                c1.font = normal_font

                c2 = ws1.cell(row=current_r, column=2, value=e_name)
                c2.alignment = Alignment(horizontal="left", vertical="center")
                c2.border = thin_border
                c2.font = bold_font

                ws1.merge_cells(start_row=current_r, start_column=3, end_row=current_r, end_column=6)
                c3 = ws1.cell(row=current_r, column=3, value=e_note)
                c3.alignment = left_align_wrap
                c3.font = normal_font
                for col_c in range(3, 7):
                    ws1.cell(row=current_r, column=col_c).border = thin_border

                current_r += 1

        ws1.column_dimensions["A"].width = 6
        ws1.column_dimensions["B"].width = 24
        ws1.column_dimensions["C"].width = 25
        ws1.column_dimensions["D"].width = 25
        ws1.column_dimensions["E"].width = 25
        ws1.column_dimensions["F"].width = 20

        # =========================================================================================
        # SHEET 2: ĐỐI SOÁT SINH HIỆU TRONG CA (SỬA LỖI GIỜ QUỐC TẾ)
        # =========================================================================================
        ws2 = wb.create_sheet(title="SINH HIỆU TRONG CA")
        ws2.views.sheetView[0].showGridLines = True

        ws2.merge_cells("A1:K1")
        ws2["A1"] = f"BẢNG ĐỐI SOÁT SINH HIỆU - {facility_name.upper()} ({shift_type_label.upper()} NGÀY {report.shift_date.strftime('%d/%m/%Y')})"
        ws2["A1"].font = title_font
        ws2["A1"].alignment = center_align
        ws2.row_dimensions[1].height = 26

        v_headers = ["STT", "Phòng", "Người Cao Tuổi", "Huyết Áp\n(mmHg)", "Mạch\n(bpm)", "SpO2\n(%)", "Nhiệt Độ\n(°C)", "Trạng Thái", "Ghi Chú Triệu Chứng", "Người Đo", "Thời Gian Đo"]
        
        ws2.row_dimensions[3].height = 25
        for c_idx, h in enumerate(v_headers, start=1):
            cell = ws2.cell(row=3, column=c_idx, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            cell.border = thin_border

        for r_i, (v, elder_name, room_num, staff_name) in enumerate(vitals, start=4):
            # Cờ bất thường (Lấy chuẩn xác từ DB do điều dưỡng tick vào)
            is_ab = bool(v.is_abnormal)
            
            # SỬA LỖI GIỜ ĐO BỊ SAI: Convert cẩn thận thời gian từ UTC -> Asia/Ho_Chi_Minh
            time_vn_str = ""
            if v.measured_at:
                dt_v_utc = pytz.utc.localize(v.measured_at) if v.measured_at.tzinfo is None else v.measured_at
                time_vn_str = dt_v_utc.astimezone(vn_tz).strftime("%H:%M:%S")

            row_vals = [
                r_i - 3,
                room_num or "N/A",
                elder_name,
                f"{v.bp_systolic}/{v.bp_diastolic}" if v.bp_systolic else "N/A",
                v.pulse or "N/A",
                f"{v.spo2}%" if v.spo2 else "N/A",
                f"{v.temperature}°C" if v.temperature else "N/A",
                "⚠️ BẤT THƯỜNG" if is_ab else "Bình thường",
                v.notes or "",
                staff_name or "N/A",
                time_vn_str
            ]
            
            ws2.row_dimensions[r_i].height = 20
            for c_i, val in enumerate(row_vals, start=1):
                c = ws2.cell(row=r_i, column=c_i, value=val)
                c.border = thin_border
                
                # Tô đỏ các ô chỉ số nếu có bất thường
                c.font = danger_font if (is_ab and c_i in [4, 5, 6, 7, 8]) else normal_font
                if is_ab: 
                    c.fill = danger_fill
                
                if c_i in [1, 2, 4, 5, 6, 7, 8, 11]:
                    c.alignment = center_align
                else:
                    c.alignment = Alignment(horizontal="left", vertical="center")

        # Độ rộng cột
        ws2.column_dimensions["A"].width = 6
        ws2.column_dimensions["B"].width = 10
        ws2.column_dimensions["C"].width = 24
        ws2.column_dimensions["D"].width = 14
        ws2.column_dimensions["E"].width = 10
        ws2.column_dimensions["F"].width = 10
        ws2.column_dimensions["G"].width = 12
        ws2.column_dimensions["H"].width = 16
        ws2.column_dimensions["I"].width = 30
        ws2.column_dimensions["J"].width = 18
        ws2.column_dimensions["K"].width = 14

        # =========================================================================================
        # 4. GÓI LẠI VÀ ĐẨY LÊN GOOGLE DRIVE
        # =========================================================================================
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        # Định dạng file theo chuẩn: BaoCaoGiaoCa_CS_1_Toi_20260815.xlsx
        filename = f"BaoCaoGiaoCa_CS_{report.facility_id}_{report.shift_type}_{report.shift_date.strftime('%Y%m%d')}_{datetime.now().strftime('%H%M%S')}.xlsx"

        drive_url = upload_shift_handover_report_to_drive(
            file_bytes=output.getvalue(),
            facility_name=facility_name,
            shift_date_str=shift_date_str,
            filename=filename
        )
        logger.info(f"[HANDOVER REPORT DRIVE SUCCESS]: Đã lưu: {filename} -> {drive_url}")

    except Exception as e:
        logger.error(f"[EXPORT REPORT FATAL ERROR]: {str(e)}")
    finally:
        db.close()