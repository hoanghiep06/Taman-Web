import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime
import shutil

# Dọn dẹp thư mục cũ để test
shutil.rmtree("Suc_Khoe_Hang_Ngay", ignore_errors=True)

def mock_append_elder_health_log(facility_name: str, elder_name: str, log_data: dict):
    # Tạo thư mục
    folder_path = f"Suc_Khoe_Hang_Ngay/{facility_name}"
    os.makedirs(folder_path, exist_ok=True)
    
    filepath = f"{folder_path}/{elder_name.replace(' ', '_')}.xlsx"
    
    if os.path.exists(filepath):
        wb = openpyxl.load_workbook(filepath)
        ws = wb.active
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Nhat Ky Suc Khoe"
        
        headers = ["Ngày Giờ", "Loại Ghi Nhận", "Huyết Áp", "Mạch", "Nhiệt độ", "SpO2", "Cân nặng", "Người Ghi Nhận", "Ghi Chú đo chỉ số", "Ghi Chú từ báo cáo giao ca"]
        ws.append(headers)
        
        header_fill = PatternFill(start_color="285A82", end_color="285A82", fill_type="solid")
        header_font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
        center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            
        ws.column_dimensions["A"].width = 16
        ws.column_dimensions["B"].width = 22
        ws.column_dimensions["C"].width = 11
        ws.column_dimensions["D"].width = 8
        ws.column_dimensions["E"].width = 10
        ws.column_dimensions["F"].width = 8
        ws.column_dimensions["G"].width = 12
        ws.column_dimensions["H"].width = 20
        ws.column_dimensions["I"].width = 35
        ws.column_dimensions["J"].width = 40

    is_update = log_data.get("is_update", False)
    date_key = log_data.get("date_key", "")
    base_type = log_data.get("base_type", "")
    target_row = ws.max_row + 1

    # TÌM DÒNG ĐỂ GHI ĐÈ NẾU LÀ CẬP NHẬT
    if is_update and date_key and base_type:
        for r_idx in range(ws.max_row, 1, -1):
            c_time = str(ws.cell(row=r_idx, column=1).value or "")
            c_type = str(ws.cell(row=r_idx, column=2).value or "")
            if date_key in c_time and base_type in c_type:
                target_row = r_idx
                break

    # Fill value
    ws.cell(row=target_row, column=1, value=log_data.get("time", ""))
    ws.cell(row=target_row, column=2, value=log_data.get("type", ""))
    ws.cell(row=target_row, column=3, value=log_data.get("bp", ""))
    ws.cell(row=target_row, column=4, value=log_data.get("pulse", ""))
    ws.cell(row=target_row, column=5, value=log_data.get("temp", ""))
    ws.cell(row=target_row, column=6, value=log_data.get("spo2", ""))
    ws.cell(row=target_row, column=7, value=log_data.get("weight", ""))
    ws.cell(row=target_row, column=8, value=log_data.get("staff", ""))
    ws.cell(row=target_row, column=9, value=log_data.get("vital_note", ""))
    ws.cell(row=target_row, column=10, value=log_data.get("shift_note", ""))
    
    danger_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
    danger_font = Font(name="Arial", size=10, color="C00000", bold=True)
    normal_font = Font(name="Arial", size=10)
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'), right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'), bottom=Side(style='thin', color='D9D9D9')
    )

    flags = log_data.get("abnormal_flags", {})

    for col_idx in range(1, 11):
        cell = ws.cell(row=target_row, column=col_idx)
        cell.border = thin_border
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        cell.font = normal_font
        cell.fill = PatternFill(fill_type=None) # Reset fill (quan trọng khi Update)

        is_danger = False
        if col_idx == 3 and flags.get("bp"): is_danger = True
        if col_idx == 4 and flags.get("pulse"): is_danger = True
        if col_idx == 5 and flags.get("temp"): is_danger = True
        if col_idx == 6 and flags.get("spo2"): is_danger = True
        if col_idx == 10 and flags.get("shift_note"): is_danger = True
        if col_idx == 9 and flags.get("any_vital"): is_danger = True

        if is_danger:
            cell.fill = danger_fill
            cell.font = danger_font

    wb.save(filepath)

# ----------------- CHẠY THỬ -----------------
fac_name = "CS1_Linh_Xuan"
date_today = "22/08/2026"

print("1. ĐO SINH HIỆU SÁNG (08:30) CHO 2 CỤ")
mock_append_elder_health_log(fac_name, "Ba Ngọc", {
    "time": f"{date_today} 08:30", "date_key": date_today, "base_type": "Sinh Hiệu",
    "type": "Đo Sinh Hiệu", "bp": "120/80", "pulse": "75", "temp": "36.5°C", "spo2": "98%", "weight": "",
    "staff": "NV Thư", "vital_note": "Bình thường", "shift_note": "",
    "is_update": False, "abnormal_flags": {"bp": False, "pulse": False, "temp": False, "spo2": False, "any_vital": False}
})

mock_append_elder_health_log(fac_name, "Má Én", {
    "time": f"{date_today} 08:35", "date_key": date_today, "base_type": "Sinh Hiệu",
    "type": "Đo Sinh Hiệu", "bp": "165/95", "pulse": "88", "temp": "38.5°C", "spo2": "96%", "weight": "",
    "staff": "NV Thư", "vital_note": "HA Cao, Sốt nhẹ", "shift_note": "",
    "is_update": False, "abnormal_flags": {"bp": True, "pulse": False, "temp": True, "spo2": False, "any_vital": True}
})

print("2. ĐO CÂN NẶNG (10:00) CHO MÁ ÉN")
mock_append_elder_health_log(fac_name, "Má Én", {
    "time": f"{date_today} 10:00", "date_key": date_today, "base_type": "Cân Nặng",
    "type": "Đo Cân Nặng", "bp": "", "pulse": "", "temp": "", "spo2": "", "weight": "42 kg",
    "staff": "NV Thư", "vital_note": "Khá gầy", "shift_note": "",
    "is_update": False, "abnormal_flags": {}
})

print("3. SỬA LẠI CÂN NẶNG (10:15) BỊ NHẬP SAI")
mock_append_elder_health_log(fac_name, "Má Én", {
    "time": f"{date_today} 10:15", "date_key": date_today, "base_type": "Cân Nặng",
    "type": "Cập nhật Cân Nặng", "bp": "", "pulse": "", "temp": "", "spo2": "", "weight": "44 kg",
    "staff": "NV Thư", "vital_note": "Sửa do nhập nhầm", "shift_note": "",
    "is_update": True, "abnormal_flags": {}
})

print("4. NỘP BÁO CÁO GIAO CA SÁNG (19:00)")
mock_append_elder_health_log(fac_name, "Má Én", {
    "time": f"{date_today} 19:00", "date_key": date_today, "base_type": "Giao ca",
    "type": "Giao ca (Ca Sáng)", "bp": "", "pulse": "", "temp": "", "spo2": "", "weight": "",
    "staff": "ĐP Tâm", "vital_note": "", "shift_note": "Sốt nguyên ngày, ca tối cần theo dõi",
    "is_update": False, "abnormal_flags": {"shift_note": True}
})

print("5. SỬA BÁO CÁO GIAO CA SÁNG (19:10)")
mock_append_elder_health_log(fac_name, "Má Én", {
    "time": f"{date_today} 19:10", "date_key": date_today, "base_type": "Giao ca",
    "type": "Giao ca (Ca Sáng) [Đã sửa]", "bp": "", "pulse": "", "temp": "", "spo2": "", "weight": "",
    "staff": "ĐP Tâm", "vital_note": "", "shift_note": "Sốt nguyên ngày, đã cho uống panadol lúc 18h",
    "is_update": True, "abnormal_flags": {"shift_note": True}
})

print("Hoàn tất! Mở thư mục Suc_Khoe_Hang_Ngay/CS1_Linh_Xuan/ để xem file.")